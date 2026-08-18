from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Q, Sum, F, ExpressionWrapper, DecimalField, Value, Case, When
from django.urls import reverse_lazy
from django.contrib import messages
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from users.mixins import ERPPermissionRequiredMixin
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
import csv
import io
from decimal import Decimal
from .models import Product, StockAdjustment, StockLedger, PerpetualCount, PerpetualCountItem
from .forms import ProductForm, StockAdjustmentForm
import json
from .services import confirm_stock_adjustment, cancel_stock_adjustment
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class ProductDetailAPIView(LoginRequiredMixin, View):
    def get(self, request, pk, *args, **kwargs):
        try:
            product = Product.objects.get(pk=pk)
            return JsonResponse({
                'id': product.id,
                'selling_price': str(product.selling_price),
                'price_tiers': [
                    {'min_quantity': tier.min_quantity, 'price': str(tier.price)}
                    for tier in product.price_tiers.order_by('-min_quantity')
                ],
                'current_stock': str(product.current_stock),
                'available_stock': str(product.available_stock),
                'reorder_level': str(product.reorder_level),
                'minimum_stock': str(product.minimum_stock),
                'allow_negative_stock': product.allow_negative_stock
            })
        except Product.DoesNotExist:
            return JsonResponse({'error': 'Product not found'}, status=404)

class ProductListView(LoginRequiredMixin, ERPPermissionRequiredMixin, ListView):
    model = Product
    template_name = 'inventory/product_list.html'
    context_object_name = 'products'
    paginate_by = 20
    permission_required = 'inventory.view_product'

    def get_queryset(self):
        qs = super().get_queryset()
        q = self.request.GET.get('q')
        category = self.request.GET.get('category')
        
        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(product_id__icontains=q))
        
        if category:
            qs = qs.filter(category=category)
            
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categories'] = Product.CategoryChoices.choices
        context['model_name'] = 'Product'
        try:
            from users.models import SavedFilter
            context['saved_filters'] = SavedFilter.objects.filter(user=self.request.user, model_name='Product')
        except ImportError:
            context['saved_filters'] = []
        return context

class PerpetualCountListView(LoginRequiredMixin, ERPPermissionRequiredMixin, ListView):
    model = PerpetualCount
    template_name = 'inventory/perpetual_counts/count_list.html'
    context_object_name = 'counts'
    paginate_by = 20
    permission_required = 'inventory.view_perpetualcount'

    def get_queryset(self):
        qs = super().get_queryset().order_by('-created_at')
        return qs

class PerpetualCountCreateView(LoginRequiredMixin, ERPPermissionRequiredMixin, View):
    permission_required = 'inventory.add_perpetualcount'
    
    def get(self, request):
        from users.models import User
        products = list(Product.objects.filter(status=True, track_stock=True).values('id', 'name', 'product_id', 'current_stock', 'stock_unit'))
        approvers = User.objects.filter(role__name='Administrator', is_active=True)
        return render(request, 'inventory/perpetual_counts/count_form.html', {
            'products_json': json.dumps(products, default=str),
            'approvers': approvers
        })
        
    def post(self, request):
        data = json.loads(request.body)
        items = data.get('items', [])
        approver_id = data.get('approver_id')
        remarks = data.get('remarks', '')
        
        if not items:
            return JsonResponse({'success': False, 'message': 'No items selected.'})
            
        count_obj = PerpetualCount.objects.create(
            status=PerpetualCount.StatusChoices.PENDING,
            created_by=request.user,
            approved_by_id=approver_id,
            remarks=remarks
        )
        
        for item in items:
            PerpetualCountItem.objects.create(
                perpetual_count=count_obj,
                product_id=item['product_id'],
                system_count=item['system_count'],
                physical_count=item['physical_count'],
                remarks=item.get('remarks', '')
            )
            
        return JsonResponse({'success': True, 'redirect_url': reverse_lazy('perpetual_count_list')})

class PerpetualCountDetailView(LoginRequiredMixin, ERPPermissionRequiredMixin, DetailView):
    model = PerpetualCount
    template_name = 'inventory/perpetual_counts/count_detail.html'
    context_object_name = 'count'
    permission_required = 'inventory.view_perpetualcount'
    
    def post(self, request, *args, **kwargs):
        count = self.get_object()
        
        if not request.user.has_perm('inventory.can_approve_perpetual_count'):
            messages.error(request, "You do not have permission to approve counts.")
            return redirect('perpetual_count_detail', pk=count.pk)
            
        action = request.POST.get('action')
        
        if count.status != PerpetualCount.StatusChoices.PENDING:
            messages.error(request, "This count has already been processed.")
            return redirect('perpetual_count_detail', pk=count.pk)
            
        if action == 'approve':
            count.status = PerpetualCount.StatusChoices.APPROVED
            count.save()
            
            # Generate adjustments
            for item in count.items.all():
                diff = item.difference
                if diff != 0:
                    adj_type = StockAdjustment.AdjustmentTypes.POSITIVE if diff > 0 else StockAdjustment.AdjustmentTypes.NEGATIVE
                    adj = StockAdjustment.objects.create(
                        date=count.date,
                        product=item.product,
                        adjustment_type=adj_type,
                        quantity=abs(diff),
                        reason='Perpetual Count Adjustment',
                        remarks=f'From {count.reference_number}',
                        status=StockAdjustment.StatusChoices.CONFIRMED,
                        created_by=request.user
                    )
                    # Create ledger entry
                    StockLedger.objects.create(
                        product=item.product,
                        tx_type=StockLedger.TransactionTypes.ADJ_POS if diff > 0 else StockLedger.TransactionTypes.ADJ_NEG,
                        qty_in=abs(diff) if diff > 0 else 0,
                        qty_out=abs(diff) if diff < 0 else 0,
                        reference_type='SYS',
                        reference_id=adj.id,
                        reference_number=adj.adjustment_number,
                        remarks=f"Perpetual Count {count.reference_number}",
                        user=request.user
                    )
                    
                    # Update cache
                    item.product.current_stock += diff
                    item.product.save(update_fields=['current_stock'])
            
            messages.success(request, f"Count {count.reference_number} approved and stock adjusted successfully.")
            
        elif action == 'reject':
            count.status = PerpetualCount.StatusChoices.REJECTED
            count.save()
            messages.warning(request, f"Count {count.reference_number} was rejected.")
            
        return redirect('perpetual_count_list')

class ProductCreateView(LoginRequiredMixin, ERPPermissionRequiredMixin, CreateView):
    model = Product
    form_class = ProductForm
    template_name = 'inventory/product_form.html'
    success_url = reverse_lazy('product_list')
    permission_required = 'inventory.add_product'

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        from .forms import ProductPriceTierFormSet
        if self.request.POST:
            data['tiers_formset'] = ProductPriceTierFormSet(self.request.POST)
        else:
            data['tiers_formset'] = ProductPriceTierFormSet()
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        tiers_formset = context['tiers_formset']
        if tiers_formset.is_valid():
            self.object = form.save()
            tiers_formset.instance = self.object
            tiers_formset.save()
            return super().form_valid(form)
        else:
            return self.render_to_response(self.get_context_data(form=form))

class ProductUpdateView(LoginRequiredMixin, ERPPermissionRequiredMixin, UpdateView):
    model = Product
    form_class = ProductForm
    template_name = 'inventory/product_form.html'
    success_url = reverse_lazy('product_list')
    permission_required = 'inventory.change_product'

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        from .forms import ProductPriceTierFormSet
        if self.request.POST:
            data['tiers_formset'] = ProductPriceTierFormSet(self.request.POST, instance=self.object)
        else:
            data['tiers_formset'] = ProductPriceTierFormSet(instance=self.object)
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        tiers_formset = context['tiers_formset']
        if tiers_formset.is_valid():
            self.object = form.save()
            tiers_formset.instance = self.object
            tiers_formset.save()
            return super().form_valid(form)
        else:
            return self.render_to_response(self.get_context_data(form=form))

class ProductDeleteView(LoginRequiredMixin, ERPPermissionRequiredMixin, DeleteView):
    model = Product
    template_name = 'inventory/product_confirm_delete.html'
    success_url = reverse_lazy('product_list')
    permission_required = 'inventory.delete_product'

    def post(self, request, *args, **kwargs):
        from django.db.models import ProtectedError
        from django.contrib import messages
        from django.shortcuts import redirect
        try:
            return super().post(request, *args, **kwargs)
        except ProtectedError:
            messages.error(request, "This product cannot be deleted because it is linked to existing records (e.g. invoices, quotations, or stock entries). Please deactivate it instead by changing its status.")
            return redirect('product_list')

def generate_products_excel(products=None, is_template=False):
    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products Export" if not is_template else "Products Import"
    
    headers = [
        'System ID', 'Product ID', 'Action', 'Name', 
        'Category', 'Brand', 'Tea Type', 'Packet Size', 'Stock Unit', 'Selling Unit',
        'Inventory Class', 'Production Type',
        'Selling Price', 'Custom Load Price',
        'Reorder Level', 'Minimum Stock', 'Track Stock', 'Allow Negative Stock', 'Tax Rate', 'Status',
        'Current Stock'
    ]
    ws.append(headers)

    column_widths = {
        'System ID': 12, 'Product ID': 15, 'Action': 12, 'Name': 35,
        'Category': 18, 'Brand': 15, 'Tea Type': 18, 'Packet Size': 15,
        'Stock Unit': 12, 'Selling Unit': 12, 'Inventory Class': 18,
        'Production Type': 22, 'Selling Price': 15, 'Custom Load Price': 18,
        'Reorder Level': 15, 'Minimum Stock': 15, 'Track Stock': 15, 'Allow Negative Stock': 22,
        'Tax Rate': 12, 'Status': 10, 'Current Stock': 15
    }

    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = column_widths.get(header, 15)

    if is_template and not products:
        ws.append([
            '', '', 'SAVE', 'Sample Product Name', 
            'Tea', 'Everbolt', 'Herbal Tea', '500g', 'pcs', 'pcs',
            'FINISHED', 'Direct Packing',
            '1500.00', '', 
            '10.00', '0.00', 'TRUE', 'FALSE', '18.00', 'TRUE',
            '100'
        ])
    elif products:
        for p in products:
            ws.append([
                str(p.id), p.product_id, 'SAVE', p.name,
                p.category,
                p.brand,
                p.tea_type or '',
                p.packet_size or '', p.stock_unit, p.selling_unit,
                p.inventory_class, p.product_type,
                str(p.selling_price), str(p.custom_load_price) if p.custom_load_price else '',
                str(p.reorder_level), str(p.minimum_stock), str(p.track_stock), str(p.allow_negative_stock), str(p.tax_rate), str(p.status),
                str(p.current_stock)
            ])

    action_dv = DataValidation(type="list", formula1='"SAVE,DELETE"', allow_blank=True)
    category_dv = DataValidation(type="list", formula1=f'"{",".join([c[0] for c in Product.CategoryChoices.choices])}"', allow_blank=True)
    brand_dv = DataValidation(type="list", formula1=f'"{",".join([c[0] for c in Product.BrandChoices.choices])}"', allow_blank=True)
    tea_type_dv = DataValidation(type="list", formula1=f'"{",".join([c[0] for c in Product.TeaTypeChoices.choices])}"', allow_blank=True)
    unit_dv = DataValidation(type="list", formula1=f'"{",".join([c[0] for c in Product.UnitTypes.choices])}"', allow_blank=True)
    inv_class_dv = DataValidation(type="list", formula1=f'"{",".join([c[0] for c in Product.InventoryClasses.choices])}"', allow_blank=True)
    prod_type_dv = DataValidation(type="list", formula1=f'"{",".join([c[0] for c in Product.ProductTypes.choices])}"', allow_blank=True)
    boolean_dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)

    # Allow custom text for these specific dropdowns
    category_dv.showErrorMessage = False
    brand_dv.showErrorMessage = False
    prod_type_dv.showErrorMessage = False

    dvs = [action_dv, category_dv, brand_dv, tea_type_dv, unit_dv, inv_class_dv, prod_type_dv, boolean_dv]
    for dv in dvs:
        ws.add_data_validation(dv)

    def apply_validation(dv, col_name):
        col_idx = headers.index(col_name) + 1
        col_letter = get_column_letter(col_idx)
        dv.add(f'{col_letter}2:{col_letter}1048576')

    apply_validation(action_dv, 'Action')
    apply_validation(category_dv, 'Category')
    apply_validation(brand_dv, 'Brand')
    apply_validation(tea_type_dv, 'Tea Type')
    apply_validation(unit_dv, 'Stock Unit')
    apply_validation(unit_dv, 'Selling Unit')
    apply_validation(inv_class_dv, 'Inventory Class')
    apply_validation(prod_type_dv, 'Production Type')
    apply_validation(boolean_dv, 'Track Stock')
    apply_validation(boolean_dv, 'Allow Negative Stock')
    apply_validation(boolean_dv, 'Status')

    return wb

@login_required
def export_all_products(request):
    products = Product.objects.all()
    q = request.GET.get('q')
    category = request.GET.get('category')
    if q:
        products = products.filter(Q(name__icontains=q) | Q(product_id__icontains=q))
    if category:
        products = products.filter(category=category)
    wb = generate_products_excel(products=products, is_template=False)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="all_products_export.xlsx"'
    wb.save(response)
    return response

@login_required
def bulk_export_products(request):
    if request.method == 'POST':
        product_ids = request.POST.getlist('selected_products')
        products = Product.objects.filter(id__in=product_ids)
        
        wb = generate_products_excel(products=products, is_template=False)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="products_export.xlsx"'
        wb.save(response)
        return response
    return redirect('product_list')

@login_required
def bulk_delete_products(request):
    if request.method == 'POST':
        product_ids = request.POST.getlist('selected_products')
        count, _ = Product.objects.filter(id__in=product_ids).delete()
        if count > 0:
            messages.success(request, f"Successfully deleted {count} products.")
    return redirect('product_list')

class ProductImportView(LoginRequiredMixin, ERPPermissionRequiredMixin, View):
    permission_required = 'inventory.add_product'

    def get(self, request):
        return render(request, 'inventory/product_import.html')

    def post(self, request):
        if 'import_file' not in request.FILES:
            messages.error(request, 'Please upload a valid Excel file.')
            return redirect('product_import')
            
        excel_file = request.FILES['import_file']
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, 'File must be an Excel (.xlsx) file.')
            return redirect('product_import')

        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
            
            created_count = 0
            updated_count = 0
            deleted_count = 0
            stock_adjusted_count = 0
            stock_errors = []   # Collect per-row stock errors to show the user
            not_found_ids = []  # Track IDs not found in DB
            
            headers = []
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(cell).strip() if cell is not None else "" for cell in row]
                else:
                    if not any(row): continue
                    row_dict = {}
                    for j, cell in enumerate(row):
                        if j < len(headers):
                            row_dict[headers[j]] = str(cell).strip() if cell is not None else ""
                    rows.append(row_dict)
            
            for row in rows:
                sys_id = row.get('System ID')
                prod_id = row.get('Product ID')
                action = row.get('Action', '').upper()
                name = row.get('Name')
                
                # Handle DELETE
                if action == 'DELETE':
                    product = None
                    if sys_id and sys_id.strip():
                        product = Product.objects.filter(id=sys_id).first()
                    elif prod_id and prod_id.strip():
                        product = Product.objects.filter(product_id=prod_id).first()
                    
                    if product:
                        product.delete()
                        deleted_count += 1
                    continue

                if not name:
                    continue
                    
                cat_name = row.get('Category')
                brand_name = row.get('Brand')
                tea_type = row.get('Tea Type')
                packet_size = row.get('Packet Size')
                stock_unit = row.get('Stock Unit')
                selling_unit = row.get('Selling Unit')
                inv_class = row.get('Inventory Class')
                prod_type = row.get('Production Type')
                
                selling_price = row.get('Selling Price')
                custom_price = row.get('Custom Load Price')
                reorder_lvl = row.get('Reorder Level')
                min_stock = row.get('Minimum Stock')
                tax_rate = row.get('Tax Rate')
                
                def parse_bool(val, default):
                    if val is None or val == '':
                        return default
                    return str(val).lower() in ['true', '1', 'yes', 'y']
                
                track_stock = parse_bool(row.get('Track Stock'), True)
                allow_neg = parse_bool(row.get('Allow Negative Stock'), False)
                status = parse_bool(row.get('Status'), True)
                
                defaults = {
                    'name': name,
                    'category': cat_name.strip() if cat_name else 'Confectionery',
                    'brand': brand_name.strip() if brand_name else 'Everbolt',
                    'tea_type': tea_type.strip() if tea_type else None,
                    'packet_size': packet_size if packet_size else None,
                    'selling_price': Decimal(selling_price.replace(',', '')) if selling_price else Decimal('0.00'),
                    'custom_load_price': Decimal(custom_price.replace(',', '')) if custom_price else None,
                    'reorder_level': Decimal(reorder_lvl.replace(',', '')) if reorder_lvl else Decimal('0.000'),
                    'minimum_stock': Decimal(min_stock.replace(',', '')) if min_stock else Decimal('0.000'),
                    'tax_rate': Decimal(tax_rate.replace(',', '')) if tax_rate else Decimal('18.00'),
                    'track_stock': track_stock,
                    'allow_negative_stock': allow_neg,
                    'status': status,
                }
                
                if stock_unit: defaults['stock_unit'] = stock_unit
                if selling_unit: defaults['selling_unit'] = selling_unit
                if inv_class: defaults['inventory_class'] = inv_class
                if prod_type: defaults['product_type'] = prod_type
                
                if prod_id: defaults['product_id'] = prod_id
                
                product = None
                is_new = False
                
                if sys_id and sys_id.strip():
                    product = Product.objects.filter(id=sys_id).first()
                if not product and prod_id:
                    product = Product.objects.filter(product_id=prod_id).first()

                if not product and (sys_id or prod_id):
                    # Product referenced by ID/Product ID but not found in DB — track it
                    not_found_ids.append(f"{name} (ID: {sys_id or prod_id})")
                    
                if product:
                    # Snapshot current_stock BEFORE the full product.save() so the
                    # stock-diff comparison below uses the DB value, not a stale cache.
                    stock_before_save = product.current_stock
                    for k, v in defaults.items():
                        setattr(product, k, v)
                    product.save()
                    updated_count += 1
                else:
                    if not_found_ids and not_found_ids[-1].startswith(name):
                        # Already tracked as not-found above, skip creating a duplicate
                        continue
                    product = Product.objects.create(**defaults)
                    created_count += 1
                    is_new = True
                    stock_before_save = product.current_stock
                
                # Handle Stock Field — Excel qty IS the physical count (the truth).
                # Logic: adjustment = Excel qty − current stock. Tally and move on.
                current_stock_val = row.get('Current Stock')
                if current_stock_val and current_stock_val.strip() and product.track_stock:
                    try:
                        qty = Decimal(current_stock_val.replace(',', ''))
                        # stock_before_save = the cached value BEFORE product.save() above
                        # (product.save() doesn't touch current_stock since it's not in defaults)
                        if stock_before_save != qty:
                            diff = qty - stock_before_save
                            tx_type = StockLedger.TransactionTypes.OPENING if (is_new or stock_before_save == 0) else (
                                StockLedger.TransactionTypes.ADJ_POS if diff > 0 else StockLedger.TransactionTypes.ADJ_NEG
                            )
                            qty_in = diff if diff > 0 else Decimal('0.0')
                            qty_out = -diff if diff < 0 else Decimal('0.0')

                            StockLedger.objects.create(
                                product=product,
                                tx_type=tx_type,
                                qty_in=qty_in,
                                qty_out=qty_out,
                                reference_type='SYS-IMPORT',
                                reference_number='IMPORT',
                                remarks='Import stock update',
                                user=request.user
                            )
                            stock_adjusted_count += 1

                        # Always set cache to Excel value — this IS the new stock count
                        product.current_stock = qty
                        product.save(update_fields=['current_stock'])
                    except Exception as ex:
                        # Capture the error with product context so the user knows what failed
                        stock_errors.append(f"{name}: {str(ex)}")
                        
            summary = f"Import successful: {created_count} created, {updated_count} updated, {deleted_count} deleted, {stock_adjusted_count} stock adjustments made."
            messages.success(request, summary)

            if not_found_ids:
                messages.warning(
                    request,
                    f"{len(not_found_ids)} product(s) in the file were not found in the database and were skipped: "
                    + ", ".join(not_found_ids)
                )

            if stock_errors:
                messages.warning(
                    request,
                    f"{len(stock_errors)} stock adjustment(s) failed: " + "; ".join(stock_errors)
                )

            return redirect('product_list')
        except Exception as e:
            messages.error(request, f"Error processing file: {str(e)}")
            return redirect('product_import')

@login_required
def download_import_template(request):
    wb = generate_products_excel(products=None, is_template=True)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="product_import_template.xlsx"'
    wb.save(response)
    return response

# Stock Adjustment Views
class StockAdjustmentListView(LoginRequiredMixin, ERPPermissionRequiredMixin, ListView):
    model = StockAdjustment
    template_name = 'inventory/adjustment_list.html'
    context_object_name = 'adjustments'
    paginate_by = 20
    ordering = ['-id']
    permission_required = 'inventory.view_stockadjustment'

class StockAdjustmentCreateView(LoginRequiredMixin, ERPPermissionRequiredMixin, CreateView):
    model = StockAdjustment
    form_class = StockAdjustmentForm
    template_name = 'inventory/adjustment_form.html'
    success_url = reverse_lazy('adjustment_list')
    permission_required = 'inventory.add_stockadjustment'

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, "Stock Adjustment created as Draft.")
        return super().form_valid(form)

class StockAdjustmentDetailView(LoginRequiredMixin, ERPPermissionRequiredMixin, DetailView):
    model = StockAdjustment
    template_name = 'inventory/adjustment_detail.html'
    context_object_name = 'adjustment'
    permission_required = 'inventory.view_stockadjustment'

from django.contrib.auth.decorators import permission_required

@login_required
@permission_required('inventory.change_stockadjustment', raise_exception=True)
def confirm_adjustment_view(request, pk):
    adjustment = get_object_or_404(StockAdjustment, pk=pk)
    if request.method == 'POST':
        try:
            confirm_stock_adjustment(adjustment, request.user)
            messages.success(request, f"Adjustment {adjustment.adjustment_number} confirmed. Stock updated.")
        except Exception as e:
            messages.error(request, f"Error: {str(e)}")
    return redirect('adjustment_list')

@login_required
@permission_required('inventory.change_stockadjustment', raise_exception=True)
def cancel_adjustment_view(request, pk):
    adjustment = get_object_or_404(StockAdjustment, pk=pk)
    if request.method == 'POST':
        try:
            cancel_stock_adjustment(adjustment, request.user)
            messages.success(request, f"Adjustment {adjustment.adjustment_number} cancelled. Stock reversed.")
        except Exception as e:
            messages.error(request, f"Error: {str(e)}")
    return redirect('adjustment_list')

# Inventory Reports/Views
# ─────────────────────────────────────────────────────────────────────────────
# STOCK SUMMARY
# ─────────────────────────────────────────────────────────────────────────────

def _apply_stock_summary_filters(qs, request):
    """Shared filter logic for StockSummaryView and StockSummaryExportView."""
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(product_id__icontains=q))

    category = request.GET.get('category', '').strip()
    if category:
        qs = qs.filter(category=category)

    inv_class = request.GET.get('inv_class', '').strip()
    if inv_class:
        qs = qs.filter(inventory_class=inv_class)

    stock_status = request.GET.get('stock_status', '').strip()
    if stock_status == 'out':
        qs = qs.filter(current_stock__lte=0)
    elif stock_status == 'low':
        qs = qs.filter(current_stock__gt=0, current_stock__lte=F('reorder_level'))
    elif stock_status == 'ok':
        qs = qs.filter(current_stock__gt=F('reorder_level'))

    return qs


class StockSummaryView(LoginRequiredMixin, ERPPermissionRequiredMixin, ListView):
    model = Product
    template_name = 'inventory/stock_summary.html'
    context_object_name = 'products'
    paginate_by = 25
    permission_required = 'inventory.view_product'

    def get_queryset(self):
        qs = Product.objects.filter(track_stock=True).select_related().order_by('category', 'name')
        return _apply_stock_summary_filters(qs, self.request)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # KPI aggregates over the FULL filtered set (not just current page)
        full_qs = self.get_queryset()
        total = full_qs.count()
        out_of_stock = full_qs.filter(current_stock__lte=0).count()
        low_stock = full_qs.filter(current_stock__gt=0, current_stock__lte=F('reorder_level')).count()
        stock_value = full_qs.aggregate(
            val=Sum(ExpressionWrapper(F('current_stock') * F('selling_price'), output_field=DecimalField()))
        )['val'] or Decimal('0.00')
        context['kpi_total'] = total
        context['kpi_out'] = out_of_stock
        context['kpi_low'] = low_stock
        context['kpi_ok'] = total - out_of_stock - low_stock
        context['kpi_value'] = stock_value
        context['categories'] = Product.CategoryChoices.choices
        context['inv_classes'] = Product.InventoryClasses.choices
        return context


class StockSummaryExportView(LoginRequiredMixin, ERPPermissionRequiredMixin, View):
    permission_required = 'inventory.view_product'

    def get(self, request, *args, **kwargs):
        qs = Product.objects.filter(track_stock=True).select_related().order_by('category', 'name')
        qs = _apply_stock_summary_filters(qs, request)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Stock Summary'

        # Styles
        header_fill = PatternFill('solid', fgColor='1E3A5F')
        header_font = Font(color='FFFFFF', bold=True, size=10)
        border = Border(
            left=Side(style='thin', color='D0D7DE'),
            right=Side(style='thin', color='D0D7DE'),
            top=Side(style='thin', color='D0D7DE'),
            bottom=Side(style='thin', color='D0D7DE'),
        )
        red_fill = PatternFill('solid', fgColor='FEE2E2')
        orange_fill = PatternFill('solid', fgColor='FEF3C7')
        green_fill = PatternFill('solid', fgColor='D1FAE5')

        headers = [
            'Product ID', 'Product Name', 'Category', 'Brand',
            'Inventory Class', 'Production Type', 'Stock Unit',
            'Current Stock', 'Reorder Level', 'Min Stock',
            'Selling Price (Ex-VAT)', 'Stock Value (Ex-VAT)', 'Status'
        ]
        ws.append(headers)
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        ws.row_dimensions[1].height = 20

        for row_num, p in enumerate(qs, 2):
            if p.current_stock <= 0:
                status = 'Out of Stock'
                row_fill = red_fill
            elif p.current_stock <= p.reorder_level:
                status = 'Low Stock'
                row_fill = orange_fill
            else:
                status = 'In Stock'
                row_fill = green_fill

            stock_value = (p.current_stock * p.selling_price).quantize(Decimal('0.01'))

            row_data = [
                p.product_id, p.name, p.get_category_display(), p.get_brand_display(),
                p.get_inventory_class_display(), p.get_product_type_display(),
                p.get_stock_unit_display(),
                float(p.current_stock), float(p.reorder_level), float(p.minimum_stock),
                float(p.selling_price), float(stock_value), status
            ]
            ws.append(row_data)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.border = border
                cell.alignment = Alignment(vertical='center')
                if col_idx in (8, 9, 10, 11, 12):
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.number_format = '#,##0.000'
                if col_idx in (11, 12):
                    cell.number_format = '#,##0.00'
                if col_idx == 13:
                    cell.fill = row_fill
                    cell.font = Font(bold=True)

        # Column widths
        col_widths = [14, 45, 18, 12, 18, 22, 12, 16, 14, 12, 20, 20, 14]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Freeze header
        ws.freeze_panes = 'A2'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        from django.utils import timezone
        ts = timezone.now().strftime('%Y%m%d_%H%M')
        response['Content-Disposition'] = f'attachment; filename="stock_summary_{ts}.xlsx"'
        return response


# ─────────────────────────────────────────────────────────────────────────────
# STOCK LEDGER
# ─────────────────────────────────────────────────────────────────────────────

def _apply_ledger_filters(qs, request):
    """Shared filter logic for StockLedgerView and StockLedgerExportView."""
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(product__product_id__icontains=q) |
            Q(product__name__icontains=q) |
            Q(reference_number__icontains=q)
        )

    tx_type = request.GET.get('tx_type', '').strip()
    if tx_type:
        qs = qs.filter(tx_type=tx_type)

    date_from = request.GET.get('date_from', '').strip()
    if date_from:
        qs = qs.filter(date__date__gte=date_from)

    date_to = request.GET.get('date_to', '').strip()
    if date_to:
        qs = qs.filter(date__date__lte=date_to)

    return qs


class StockLedgerView(LoginRequiredMixin, ERPPermissionRequiredMixin, ListView):
    model = StockLedger
    template_name = 'inventory/stock_ledger.html'
    context_object_name = 'entries'
    paginate_by = 30
    ordering = ['-date']
    permission_required = 'inventory.view_product'

    def get_queryset(self):
        qs = super().get_queryset().select_related('product', 'user')
        return _apply_ledger_filters(qs, self.request)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        full_qs = self.get_queryset()
        agg = full_qs.aggregate(
            total_in=Sum('qty_in'),
            total_out=Sum('qty_out'),
            count=Sum(Value(1))
        )
        total_in = agg['total_in'] or Decimal('0')
        total_out = agg['total_out'] or Decimal('0')
        context['kpi_in'] = total_in
        context['kpi_out'] = total_out
        context['kpi_net'] = total_in - total_out
        context['kpi_count'] = full_qs.count()
        context['tx_type_choices'] = StockLedger.TransactionTypes.choices
        return context


class StockLedgerExportView(LoginRequiredMixin, ERPPermissionRequiredMixin, View):
    permission_required = 'inventory.view_product'

    def get(self, request, *args, **kwargs):
        qs = StockLedger.objects.select_related('product', 'user').order_by('-date')
        qs = _apply_ledger_filters(qs, request)

        # Limit to 10,000 rows for safety
        qs = qs[:10000]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Stock Ledger'

        header_fill = PatternFill('solid', fgColor='1E3A5F')
        header_font = Font(color='FFFFFF', bold=True, size=10)
        border = Border(
            left=Side(style='thin', color='D0D7DE'),
            right=Side(style='thin', color='D0D7DE'),
            top=Side(style='thin', color='D0D7DE'),
            bottom=Side(style='thin', color='D0D7DE'),
        )

        TX_COLORS = {
            'OPENING':  'EDE9FE',
            'GRN':      'D1FAE5',
            'PROD_CONS':'FEF3C7',
            'PROD_OUT': 'D1FAE5',
            'SALES_ISS':'FEE2E2',
            'SALES_RET':'D1FAE5',
            'PURC_RET': 'D1FAE5',
            'ADJ_POS':  'DBEAFE',
            'ADJ_NEG':  'FEE2E2',
        }

        headers = [
            'Date', 'Time', 'Product ID', 'Product Name',
            'Transaction Type', 'Reference Type', 'Reference Number',
            'IN (+)', 'OUT (-)', 'Remarks', 'User'
        ]
        ws.append(headers)
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        ws.row_dimensions[1].height = 20

        for row_num, e in enumerate(qs, 2):
            row_fill = PatternFill('solid', fgColor=TX_COLORS.get(e.tx_type, 'FFFFFF'))
            row_data = [
                e.date.strftime('%Y-%m-%d'),
                e.date.strftime('%H:%M:%S'),
                e.product.product_id,
                e.product.name,
                e.get_tx_type_display(),
                e.reference_type,
                e.reference_number,
                float(e.qty_in) if e.qty_in else 0,
                float(e.qty_out) if e.qty_out else 0,
                e.remarks or '',
                e.user.get_full_name() if e.user else ''
            ]
            ws.append(row_data)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.border = border
                cell.fill = row_fill
                cell.alignment = Alignment(vertical='center')
                if col_idx in (8, 9):
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.number_format = '#,##0.000'
                    if col_idx == 8 and float(e.qty_in or 0) > 0:
                        cell.font = Font(color='15803D', bold=True)
                    elif col_idx == 9 and float(e.qty_out or 0) > 0:
                        cell.font = Font(color='B91C1C', bold=True)

        col_widths = [12, 10, 16, 45, 26, 16, 24, 14, 14, 40, 20]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = 'A2'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        from django.utils import timezone
        ts = timezone.now().strftime('%Y%m%d_%H%M')
        response['Content-Disposition'] = f'attachment; filename="stock_ledger_{ts}.xlsx"'
        return response
