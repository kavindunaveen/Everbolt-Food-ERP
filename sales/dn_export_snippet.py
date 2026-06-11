class DeliveryNoteExportView(LoginRequiredMixin, ERPPermissionRequiredMixin, View):
    permission_required = 'sales.view_deliverynote'
    
    def get(self, request, *args, **kwargs):
        import openpyxl
        from openpyxl.utils import get_column_letter

        qs = DeliveryNote.objects.all().order_by('-created_at')
        
        # Status Filter
        status = request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        
        # Delivered By Filter
        delivered_by = request.GET.get('delivered_by')
        if delivered_by:
            qs = qs.filter(delivered_by_id=delivered_by)
            
        # Date Range Filter
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        # Delivery Date Range Filter
        delivery_date_from = request.GET.get('delivery_date_from')
        delivery_date_to = request.GET.get('delivery_date_to')
        if delivery_date_from:
            qs = qs.filter(delivery_date__gte=delivery_date_from)
        if delivery_date_to:
            qs = qs.filter(delivery_date__lte=delivery_date_to)

        # Unified Search
        q = request.GET.get('q')
        if q:
            from django.db.models import Q
            qs = qs.filter(
                Q(dn_number__icontains=q) |
                Q(invoice__invoice_number__icontains=q) |
                Q(customer_name__icontains=q)
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Delivery Notes Export"
        
        headers = [
            'DN Number', 'Invoice Number', 'Customer', 'Delivery Address', 
            'Target Delivery Date', 'Delivered By', 'Status', 'Created At', 'Remarks'
        ]
        ws.append(headers)

        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
            
        for dn in qs:
            ws.append([
                dn.dn_number,
                dn.invoice.invoice_number,
                dn.customer_name,
                dn.delivery_address,
                dn.delivery_date.strftime('%Y-%m-%d') if dn.delivery_date else '',
                dn.delivery_officer_name,
                dn.get_status_display(),
                dn.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                dn.remarks or ''
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="delivery_notes.xlsx"'
        wb.save(response)
        return response
