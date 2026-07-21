from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import TemplateView, ListView, CreateView, UpdateView, View, DetailView, DeleteView
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from users.mixins import ERPPermissionRequiredMixin
from django.contrib.auth.decorators import login_required, permission_required
from django.core.exceptions import ValidationError
from django.contrib import messages
from django.urls import reverse_lazy, reverse
from django.db import transaction
from django.utils import timezone
from django.db.models import Sum, F
from decimal import Decimal, ROUND_UP, ROUND_HALF_UP
from .models import Quotation, Invoice, DeliveryNote, DeliveryNoteItem, SalesAuditLog, Return, CreditNote
from .forms import QuotationForm, QuotationItemFormSet, InvoiceForm, InvoiceItemFormSet, DeliveryNoteForm
from .services import (
    issue_invoice, cancel_invoice, send_invoice_approval_email, process_return,
    deduct_dn_stock, restore_dn_stock, log_sales_event, update_stock_reserves
)
from users.models import SavedFilter
from django.contrib.contenttypes.models import ContentType
import csv
import json
from datetime import timedelta
from django.db.models.functions import TruncDate
from num2words import num2words


from users.mixins import ERPUserPassesTestMixin
class AdminRequiredMixin(ERPUserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_admin()

class MainDashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'sales/main_dashboard.html'

class SalesDashboardView(LoginRequiredMixin, ERPPermissionRequiredMixin, TemplateView):
    template_name = 'sales/dashboard.html'
    permission_required = 'sales.view_invoice'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from users.models import User
        from django.db.models import Q
        context['sales_officers'] = User.objects.filter(role__name='Sales Officer', is_active=True).distinct()
        context['model_name'] = 'SalesDashboard'
        
        q = self.request.GET.get('q')
        
        today = timezone.now().date()
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        
        all_time = self.request.GET.get('all_time') == 'true'
        
        if all_time:
            date_from = None
            date_to = None
        elif not date_from and not date_to:
            date_from = today.strftime('%Y-%m-%d')
            date_to = today.strftime('%Y-%m-%d')
            
        salesperson_id = self.request.GET.get('salesperson')
        
        # Generate quick_months for Quick Select
        import calendar
        quick_months = []
        today_str = today.strftime('%Y-%m-%d')
        quick_months.append({
            'label': 'Today',
            'date_from': today_str,
            'date_to': today_str,
        })
        yesterday_str = (today - timezone.timedelta(days=1)).strftime('%Y-%m-%d')
        quick_months.append({
            'label': 'Yesterday',
            'date_from': yesterday_str,
            'date_to': yesterday_str,
        })
        for i in range(5):
            first_day = (today.replace(day=1) - timezone.timedelta(days=30 * i)).replace(day=1)
            last_day = calendar.monthrange(first_day.year, first_day.month)[1]
            last_date = first_day.replace(day=last_day)
            
            label = first_day.strftime('%b %Y')
            if i == 0:
                label = 'Current Month'
            elif i == 1:
                label = 'Last Month'

            quick_months.append({
                'label': label,
                'date_from': first_day.strftime('%Y-%m-%d'),
                'date_to': last_date.strftime('%Y-%m-%d'),
            })
        context['quick_months'] = quick_months
        
        # Pass the effective dates to context so the template knows what is active
        context['active_date_from'] = date_from
        context['active_date_to'] = date_to
        
        quotations = Quotation.objects.all()
        invoices = Invoice.objects.all()
        q = self.request.GET.get('q')
        status = self.request.GET.get('status')
        
        if q:
            quotations = quotations.filter(quotation_number__icontains=q)
            invoices = invoices.filter(invoice_number__icontains=q)
            
        if status:
            quotations = quotations.filter(status=status)
            invoices = invoices.filter(status=status)
            
        if date_from:
            quotations = quotations.filter(creation_date__date__gte=date_from)
            invoices = invoices.filter(creation_date__date__gte=date_from)
            
        if date_to:
            quotations = quotations.filter(creation_date__date__lte=date_to)
            invoices = invoices.filter(creation_date__date__lte=date_to)
            
        if salesperson_id:
            quotations = quotations.filter(salesperson_id=salesperson_id)
            invoices = invoices.filter(salesperson_id=salesperson_id)
            
        context['total_quotations_count'] = quotations.count()
        context['recent_quotations'] = quotations.order_by('-creation_date')[:15]
        context['recent_invoices'] = invoices.order_by('-creation_date')[:15]
        
        # Only include confirmed sales (ISSUED or PAID) for revenue stats
        active_invoices = invoices.filter(status__in=['ISSUED', 'PAID'])
        context['total_invoice_count'] = active_invoices.count()
        
        # Calculate gross revenue - use total_amount - tax_amount to correctly include converted quotation invoices
        revenue_with_vat = active_invoices.aggregate(Sum('total_amount'))['total_amount__sum'] or Decimal('0.00')
        revenue_ex_vat = sum(
            (inv.total_amount - inv.tax_amount) for inv in active_invoices.only('total_amount', 'tax_amount')
        )
        
        # Subtract Credit Notes (Returns)
        # Note: We filter credit notes by the same salesperson/date filters if applied to invoices
        credit_notes = CreditNote.objects.filter(original_invoice__in=active_invoices)
        total_credit = credit_notes.aggregate(Sum('items__credit_amount'))['items__credit_amount__sum'] or Decimal('0.00')
        
        # We can just use total_credit for both ex-vat and with-vat estimations if tax isn't explicitly split on CN, 
        # or we can assume credit_amount is Ex-VAT and calculate accordingly. Currently we just use the raw credit_amount.
        credit_subtotal = total_credit
        
        total_revenue_with_vat = revenue_with_vat - total_credit
        total_revenue_ex_vat = revenue_ex_vat - Decimal(str(credit_subtotal))
        
        context['total_revenue_with_vat'] = total_revenue_with_vat
        context['total_revenue_ex_vat'] = total_revenue_ex_vat
        
        # Calculate Average Daily Sales
        from django.db.models.functions import TruncDate
        days_recorded = active_invoices.annotate(d=TruncDate('creation_date')).values('d').distinct().count()
        if days_recorded == 0:
            days_recorded = 1
        
        context['days_recorded'] = days_recorded
        context['avg_daily_sales'] = float(total_revenue_ex_vat) / days_recorded
        
        sales_officers = context['sales_officers']
        
        try:
            from users.models import SavedFilter
            context['saved_filters'] = SavedFilter.objects.filter(user=self.request.user, model_name='SalesDashboard')
        except ImportError:
            context['saved_filters'] = []
            
        # Determine target month/year from date_from or today
        target_date = timezone.now().date()
        if date_from:
            from datetime import datetime
            if isinstance(date_from, str):
                try:
                    target_date = datetime.strptime(date_from, '%Y-%m-%d').date()
                except ValueError:
                    pass
            else:
                target_date = date_from

        target_year = target_date.year
        target_month = target_date.month

        # Fetch Salesperson targets for this month
        from dashboard.models import SalespersonTarget
        sp_targets_map = {}
        for spt in SalespersonTarget.objects.filter(year=target_year, month=target_month):
            sp_targets_map[spt.salesperson_id] = spt.target_value

        # Calculate performance per Sales Officer
        officer_performance = []
        for officer in sales_officers:
            officer_invs = active_invoices.filter(salesperson=officer)
            officer_cns = credit_notes.filter(original_invoice__in=officer_invs)
            
            off_rev_with = officer_invs.aggregate(Sum('total_amount'))['total_amount__sum'] or Decimal('0.00')
            off_rev_ex = sum(
                (inv.total_amount - inv.tax_amount) for inv in officer_invs.only('total_amount', 'tax_amount')
            )
            
            off_cred = officer_cns.aggregate(Sum('items__credit_amount'))['items__credit_amount__sum'] or Decimal('0.00')
            off_cred_sub = off_cred
            
            net_with = off_rev_with - off_cred
            net_ex = Decimal(str(off_rev_ex)) - Decimal(str(off_cred_sub))

            
            sp_target_val = sp_targets_map.get(officer.id, Decimal('0.00'))
            
            target = sp_target_val
            if target > 0:
                progress_pct = min((net_ex / target) * 100, Decimal('100.00'))
            else:
                progress_pct = Decimal('0.00')
                
            if total_revenue_ex_vat > 0:
                contribution_pct = (net_ex / total_revenue_ex_vat) * Decimal('100')
            else:
                contribution_pct = Decimal('0.00')
                
            officer_performance.append({
                'officer': officer,
                'total_sales': net_with,
                'total_ex_vat': net_ex,
                'invoice_count': officer_invs.count(),
                'target': target,
                'progress_pct': progress_pct,
                'contribution_pct': contribution_pct
            })
        
        # Sort by total_ex_vat descending for leaderboard
        officer_performance.sort(key=lambda x: x['total_ex_vat'], reverse=True)
        context['officer_performance'] = officer_performance
        
        # --- Chart Data Generation ---
        
        # 1. Leaderboard Data (Bar Chart)
        # Using the sorted officer_performance list
        leaderboard_names = [perf['officer'].get_full_name() or perf['officer'].username for perf in officer_performance]
        leaderboard_sales = [float(perf['total_ex_vat']) for perf in officer_performance]
        leaderboard_targets = [float(perf['target']) for perf in officer_performance]
        context['leaderboard_names_json'] = json.dumps(leaderboard_names)
        context['leaderboard_sales_json'] = json.dumps(leaderboard_sales)
        context['leaderboard_targets_json'] = json.dumps(leaderboard_targets)
        context['target_month_name'] = target_date.strftime('%B %Y')
        
        # 1.5 Contribution Data (Doughnut Chart)
        contribution_names = [perf['officer'].get_full_name() or perf['officer'].username for perf in officer_performance if perf['total_ex_vat'] > 0]
        contribution_sales = [float(perf['total_ex_vat']) for perf in officer_performance if perf['total_ex_vat'] > 0]
        context['contribution_names_json'] = json.dumps(contribution_names)
        context['contribution_sales_json'] = json.dumps(contribution_sales)
        
        # 2. Revenue Trendline Data (Line Chart)
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=29)
        
        trend_invoices = active_invoices
        if not date_from and not date_to:
            trend_invoices = trend_invoices.filter(creation_date__date__gte=start_date)
            
        daily_revenue = trend_invoices.annotate(
            date=TruncDate('creation_date'),
            ex_vat=F('total_amount') - F('tax_amount')
        ).values('date').annotate(
            daily_total=Sum('ex_vat')
        ).order_by('date')
        
        # To handle credit notes in trend, we can just map invoice creation date to revenue.
        # This is an approximation for trend line.
        revenue_dict = {str(item['date']): float(item['daily_total']) for item in daily_revenue}
        
        trend_dates = []
        trend_totals = []
        
        if date_from and date_to:
            # If explicit filters are set, just plot the available data points
            for item in daily_revenue:
                trend_dates.append(item['date'].strftime('%b %d'))
                trend_totals.append(float(item['daily_total']))
        else:
            # Default 30 days smooth line
            curr = start_date
            while curr <= end_date:
                trend_dates.append(curr.strftime('%b %d'))
                trend_totals.append(revenue_dict.get(str(curr), 0.0))
                curr += timedelta(days=1)
                
        context['trend_dates_json'] = json.dumps(trend_dates)
        context['trend_totals_json'] = json.dumps(trend_totals)
            
        return context

class QuotationListView(LoginRequiredMixin, ERPPermissionRequiredMixin, ListView):
    model = Quotation
    template_name = 'sales/quotation_list.html'
    context_object_name = 'quotations'
    paginate_by = 20
    permission_required = 'sales.view_quotation'
    
    def get_queryset(self):
        from django.db.models import Q
        qs = super().get_queryset().select_related('customer', 'salesperson').order_by('-creation_date')

        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)

        date_from = self.request.GET.get('date_from')
        if date_from:
            qs = qs.filter(creation_date__date__gte=date_from)

        date_to = self.request.GET.get('date_to')
        if date_to:
            qs = qs.filter(creation_date__date__lte=date_to)

        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(
                Q(quotation_number__icontains=q) |
                Q(customer__customer_name__icontains=q) |
                Q(customer__customer_code__icontains=q)
            )

        salesperson_id = self.request.GET.get('salesperson')
        if salesperson_id:
            qs = qs.filter(salesperson_id=salesperson_id)

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from users.models import User
        from django.db.models import Q
        context['sales_officers'] = User.objects.filter(role__name='Sales Officer', is_active=True).distinct()
        try:
            from users.models import SavedFilter
            context['saved_filters'] = SavedFilter.objects.filter(user=self.request.user, model_name='Quotation')
        except ImportError:
            context['saved_filters'] = []
        context['model_name'] = 'Quotation'
        return context

class InvoiceListView(LoginRequiredMixin, ERPPermissionRequiredMixin, ListView):
    model = Invoice
    template_name = 'sales/invoice_list.html'
    context_object_name = 'invoices'
    paginate_by = 20
    permission_required = 'sales.view_invoice'
    
    def get_queryset(self):
        from django.db.models import Q
        qs = super().get_queryset().select_related('customer', 'salesperson').prefetch_related('delivery_notes').order_by('-creation_date')

        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)

        date_from = self.request.GET.get('date_from')
        if date_from:
            qs = qs.filter(creation_date__date__gte=date_from)

        date_to = self.request.GET.get('date_to')
        if date_to:
            qs = qs.filter(creation_date__date__lte=date_to)

        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(
                Q(invoice_number__icontains=q) |
                Q(customer__customer_name__icontains=q) |
                Q(customer__customer_code__icontains=q)
            )

        salesperson_id = self.request.GET.get('salesperson')
        if salesperson_id:
            qs = qs.filter(salesperson_id=salesperson_id)

        is_returned = self.request.GET.get('is_returned')
        if is_returned == 'true':
            qs = qs.filter(status__in=['CANCELLED', 'CANCEL_PENDING'], cancellation_reason__icontains='Customer Return')

        dn_status = self.request.GET.get('dn_status')
        if dn_status == 'has_dn':
            qs = qs.filter(delivery_notes__isnull=False).distinct()
        elif dn_status == 'pending_dn':
            qs = qs.filter(status='ISSUED', delivery_notes__isnull=True).distinct()

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from users.models import User
        from django.db.models import Q
        context['sales_officers'] = User.objects.filter(role__name='Sales Officer', is_active=True).distinct()
        try:
            from users.models import SavedFilter
            context['saved_filters'] = SavedFilter.objects.filter(user=self.request.user, model_name='Invoice')
        except ImportError:
            context['saved_filters'] = []
        context['model_name'] = 'Invoice'
        return context

class QuotationCreateView(LoginRequiredMixin, ERPPermissionRequiredMixin, CreateView):
    model = Quotation
    form_class = QuotationForm
    template_name = 'sales/quotation_form.html'
    success_url = reverse_lazy('quotation_list')
    permission_required = 'sales.add_quotation'

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['items'] = QuotationItemFormSet(self.request.POST)
        else:
            data['items'] = QuotationItemFormSet()
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        items = context['items']
        with transaction.atomic():
            self.object = form.save(commit=False)
            self.object.salesperson = self.object.customer.assigned_sales_officer or self.request.user
            self.object.created_by = self.request.user
            # Quotation number is generated automatically in Quotation.save()
            
            if items.is_valid():
                self.object.save()
                items.instance = self.object
                
                saved_items = items.save(commit=False)
                
                total = 0
                tax = 0
                tot_discount = 0
                for item in saved_items:
                    item.quotation = self.object
                    discount_amt = item.get_discount_amount
                    if self.object.customer.vat_enabled:
                        item.tax_amount = ((item.quantity * item.unit_price) - discount_amt) * Decimal('0.18')
                    else:
                        item.tax_amount = Decimal('0.00')
                        
                    item.line_total = (item.quantity * item.unit_price) - discount_amt + item.tax_amount
                    item.save()
                    
                for obj in items.deleted_objects:
                    obj.delete()
                    
                # Calculate aggregated values for quotation
                from .models import QuotationItem
                current_items = QuotationItem.objects.filter(quotation=self.object)
                gross_total = sum((item.quantity * item.unit_price) for item in current_items)
                line_discount = sum(item.get_discount_amount for item in current_items)
                subtotal = gross_total - line_discount
                
                custom_val = self.object.custom_discount_value or Decimal('0.00')
                if self.object.custom_discount_type == 'PERCENT':
                    global_discount = subtotal * (custom_val / Decimal('100.0'))
                else:
                    global_discount = custom_val
                    
                tot_discount = line_discount + global_discount
                subtotal -= global_discount
                if subtotal < Decimal('0.00'):
                    subtotal = Decimal('0.00')
                    
                if self.object.customer.vat_enabled:
                    tax = subtotal * Decimal('0.18')
                else:
                    tax = Decimal('0.00')
                
                total = subtotal + tax
                
                self.object.tax_amount = tax
                self.object.subtotal_amount = subtotal
                self.object.total_discount = tot_discount
                self.object.total_amount = total.quantize(Decimal('1.'), rounding=ROUND_HALF_UP)
                self.object.save()
            else:
                return super().form_invalid(form)
            
            
        log_sales_event(self.object, self.request.user, "Quotation Created", new_value=self.object.get_status_display())
        return super().form_valid(form)

class QuotationUpdateView(LoginRequiredMixin, ERPPermissionRequiredMixin, UpdateView):
    model = Quotation
    form_class = QuotationForm
    template_name = 'sales/quotation_form.html'
    success_url = reverse_lazy('quotation_list')
    permission_required = 'sales.change_quotation'

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['items'] = QuotationItemFormSet(self.request.POST, instance=self.object)
        else:
            data['items'] = QuotationItemFormSet(instance=self.object)
        
        from .models import SalesAuditLog
        from django.contrib.contenttypes.models import ContentType
        ct = ContentType.objects.get_for_model(Quotation)
        data['audit_logs'] = SalesAuditLog.objects.filter(content_type=ct, object_id=self.object.id).order_by('-timestamp')
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        items = context['items']
        with transaction.atomic():
            self.object = form.save(commit=False)
            if self.object.status == 'DRAFT':
                self.object.salesperson = self.object.customer.assigned_sales_officer or self.request.user
            self.object.save()
            if items.is_valid():
                items.instance = self.object
                saved_items = items.save(commit=False)
                
                total = 0
                tax = 0
                tot_discount = 0
                for item in saved_items:
                    item.quotation = self.object
                    discount_amt = item.get_discount_amount
                    if self.object.customer.vat_enabled:
                        item.tax_amount = ((item.quantity * item.unit_price) - discount_amt) * Decimal('0.18')
                    else:
                        item.tax_amount = Decimal('0.00')
                        
                    item.line_total = (item.quantity * item.unit_price) - discount_amt + item.tax_amount
                    item.save()
                    
                for obj in items.deleted_objects:
                    obj.delete()
                    
                # Re-calculate totals from ALL items associated with this quotation
                from .models import QuotationItem
                current_items = QuotationItem.objects.filter(quotation=self.object)
                gross_total = sum((item.quantity * item.unit_price) for item in current_items)
                line_discount = sum(item.get_discount_amount for item in current_items)
                subtotal = gross_total - line_discount
                
                custom_val = self.object.custom_discount_value or Decimal('0.00')
                if self.object.custom_discount_type == 'PERCENT':
                    global_discount = subtotal * (custom_val / Decimal('100.0'))
                else:
                    global_discount = custom_val
                    
                tot_discount = line_discount + global_discount
                subtotal -= global_discount
                if subtotal < Decimal('0.00'):
                    subtotal = Decimal('0.00')
                    
                if self.object.customer.vat_enabled:
                    tax = subtotal * Decimal('0.18')
                else:
                    tax = Decimal('0.00')
                
                total = subtotal + tax
                
                self.object.tax_amount = tax
                self.object.subtotal_amount = subtotal
                self.object.total_discount = tot_discount
                self.object.total_amount = total.quantize(Decimal('1.'), rounding=ROUND_HALF_UP)
                self.object.save()
            else:
                return super().form_invalid(form)
            
        return super().form_valid(form)


class InvoiceCreateView(LoginRequiredMixin, ERPPermissionRequiredMixin, CreateView):
    model = Invoice
    form_class = InvoiceForm
    template_name = 'sales/invoice_form.html'
    success_url = reverse_lazy('invoice_list')
    permission_required = 'sales.add_invoice'

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['items'] = InvoiceItemFormSet(self.request.POST)
        else:
            data['items'] = InvoiceItemFormSet()
            
        from users.models import User
        # Retrieve all active users who legitimately have permission to approve invoices.
        approving_users = [u for u in User.objects.filter(is_active=True) if u.has_perm('sales.approve_invoice') and u != self.request.user]
        data['approvers'] = approving_users
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        items = context['items']
        with transaction.atomic():
            self.object = form.save(commit=False)
            self.object.salesperson = self.object.customer.assigned_sales_officer or self.request.user
            self.object.created_by = self.request.user
            # Invoice number is generated automatically in Invoice.save()

            # ── Snapshot the selected delivery address ──────────────────────────
            self.object.snap_delivery_line1    = self.request.POST.get('snap_delivery_line1', '').strip() or None
            self.object.snap_delivery_line2    = self.request.POST.get('snap_delivery_line2', '').strip() or None
            self.object.snap_delivery_city     = self.request.POST.get('snap_delivery_city', '').strip() or None
            self.object.snap_delivery_province = self.request.POST.get('snap_delivery_province', '').strip() or None
            self.object.snap_delivery_zip      = self.request.POST.get('snap_delivery_zip', '').strip() or None
            # Fallback to customer's old single delivery address if nothing posted
            if not any([self.object.snap_delivery_line1, self.object.snap_delivery_city]):
                c = self.object.customer
                self.object.snap_delivery_line1    = c.delivery_address_line1
                self.object.snap_delivery_line2    = c.delivery_address_line2
                self.object.snap_delivery_city     = c.delivery_city
                self.object.snap_delivery_province = c.delivery_province
                self.object.snap_delivery_zip      = c.delivery_zip_code
            
            # ── Block based on customer status & minimum stock ──
            if items.is_valid():
                for form_item in items:
                    if form_item.cleaned_data and not form_item.cleaned_data.get('DELETE', False):
                        product = form_item.cleaned_data.get('product')
                        quantity = form_item.cleaned_data.get('quantity')
                        if product and quantity:
                            if product.minimum_stock > 0 and (product.available_stock - quantity) < product.minimum_stock:
                                form.add_error(None, ValidationError(f"Cannot invoice: Stock for {product.name} will drop below the Minimum Stock limit."))
                                return super().form_invalid(form)
                            if not product.allow_negative_stock and quantity > product.available_stock:
                                form.add_error(None, ValidationError(f"Cannot invoice: Insufficient available stock for {product.name}."))
                                return super().form_invalid(form)

            requires_approval = (self.object.customer.customer_status in ['BLACKLIST', 'ONHOLD'])

            if requires_approval and not getattr(self.object, 'is_approved', False):
                if self.request.POST.get('is_approval_request') == 'true':
                    self.object.status = 'APPROVAL_PENDING'
                    approver_id = self.request.POST.get('designated_approver')
                    if approver_id:
                        from users.models import User
                        try:
                            self.object.designated_approver = User.objects.get(pk=approver_id)
                        except User.DoesNotExist:
                            pass
                else:
                    form.add_error(None, ValidationError(f"Invoice cannot be saved because customer is {self.object.customer.customer_status}."))
                    return super().form_invalid(form)
            
            if items.is_valid():
                self.object.save()
                items.instance = self.object
                
                saved_items = items.save(commit=False)
                
                total = 0
                tax = 0
                tot_discount = 0
                for item in saved_items:
                    item.invoice = self.object
                    discount_amt = item.get_discount_amount
                    if self.object.customer.vat_enabled:
                        item.tax_amount = ((item.quantity * item.unit_price) - discount_amt) * Decimal('0.18')
                    else:
                        item.tax_amount = Decimal('0.00')
                        
                    item.line_total = (item.quantity * item.unit_price) - discount_amt + item.tax_amount
                    item.save()
                    
                for obj in items.deleted_objects:
                    obj.delete()
                
                # Calculate aggregated values for invoice
                from .models import InvoiceItem
                current_items = InvoiceItem.objects.filter(invoice=self.object)
                gross_total = sum((item.quantity * item.unit_price) for item in current_items)
                line_discount = sum(item.get_discount_amount for item in current_items)
                subtotal = gross_total - line_discount
                
                custom_val = self.object.custom_discount_value or Decimal('0.00')
                if self.object.custom_discount_type == 'PERCENT':
                    global_discount = subtotal * (custom_val / Decimal('100.0'))
                else:
                    global_discount = custom_val
                    
                tot_discount = line_discount + global_discount
                subtotal -= global_discount
                if subtotal < Decimal('0.00'):
                    subtotal = Decimal('0.00')
                    
                if self.object.customer.vat_enabled:
                    tax = subtotal * Decimal('0.18')
                else:
                    tax = Decimal('0.00')
                
                total = subtotal + tax
                
                self.object.tax_amount = tax
                self.object.subtotal_amount = subtotal
                self.object.total_discount = tot_discount
                self.object.total_amount = total.quantize(Decimal('1.'), rounding=ROUND_HALF_UP)
                self.object.save()
                
                log_sales_event(
                    obj=self.object,
                    user=self.request.user,
                    action="Invoice Created",
                    new_value=self.object.get_status_display(),
                    notes=f"Initial creation. Approver: {self.object.designated_approver}" if self.object.status == 'APPROVAL_PENDING' else None
                )
                
                update_stock_reserves(self.object)
                
                if getattr(self.object, 'status', None) == 'APPROVAL_PENDING':
                    send_invoice_approval_email(self.object, self.request)
            else:
                return super().form_invalid(form)
            
        return super().form_valid(form)

class InvoiceUpdateView(LoginRequiredMixin, ERPPermissionRequiredMixin, UpdateView):
    model = Invoice
    form_class = InvoiceForm
    template_name = 'sales/invoice_form.html'
    success_url = reverse_lazy('invoice_list')
    permission_required = 'sales.change_invoice'

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['items'] = InvoiceItemFormSet(self.request.POST, instance=self.object)
        else:
            data['items'] = InvoiceItemFormSet(instance=self.object)
            
        from users.models import User
        approving_users = [u for u in User.objects.filter(is_active=True) if u.has_perm('sales.approve_invoice') and u != self.request.user]
        data['approvers'] = approving_users
        
        from .models import SalesAuditLog
        from django.contrib.contenttypes.models import ContentType
        ct = ContentType.objects.get_for_model(Invoice)
        data['audit_logs'] = SalesAuditLog.objects.filter(content_type=ct, object_id=self.object.id).order_by('-timestamp')
        
        # Add cancellation approvers based on permission
        data['cancellation_approvers'] = [u for u in User.objects.filter(is_active=True) if u.has_perm('sales.approve_invoice')]
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        items = context['items']
        with transaction.atomic():
            self.object = form.save(commit=False)

            # Sync salesperson for DRAFT invoices
            if self.object.status == 'DRAFT':
                self.object.salesperson = self.object.customer.assigned_sales_officer or self.request.user
                # Allow updating snapshot address on DRAFT invoices
                if self.request.POST.get('snap_delivery_line1') or self.request.POST.get('snap_delivery_city'):
                    self.object.snap_delivery_line1    = self.request.POST.get('snap_delivery_line1', '').strip() or None
                    self.object.snap_delivery_line2    = self.request.POST.get('snap_delivery_line2', '').strip() or None
                    self.object.snap_delivery_city     = self.request.POST.get('snap_delivery_city', '').strip() or None
                    self.object.snap_delivery_province = self.request.POST.get('snap_delivery_province', '').strip() or None
                    self.object.snap_delivery_zip      = self.request.POST.get('snap_delivery_zip', '').strip() or None
            
            if self.object.pk and self.object.status != 'DRAFT':
                form.add_error(None, ValidationError("Only DRAFT invoices can be edited and saved."))
                return super().form_invalid(form)
            
            # ── Block based on customer status & minimum stock ──
            if self.object.status == 'DRAFT' and items.is_valid():
                for form_item in items:
                    if form_item.cleaned_data and not form_item.cleaned_data.get('DELETE', False):
                        product = form_item.cleaned_data.get('product')
                        quantity = form_item.cleaned_data.get('quantity')
                        if product and quantity:
                            if product.minimum_stock > 0 and (product.available_stock - quantity) < product.minimum_stock:
                                form.add_error(None, ValidationError(f"Cannot invoice: Stock for {product.name} will drop below the Minimum Stock limit."))
                                return super().form_invalid(form)
                            if not product.allow_negative_stock and quantity > product.available_stock:
                                form.add_error(None, ValidationError(f"Cannot invoice: Insufficient available stock for {product.name}."))
                                return super().form_invalid(form)

            requires_approval = (self.object.customer.customer_status in ['BLACKLIST', 'ONHOLD'])

            if requires_approval and self.object.status == 'DRAFT' and not getattr(self.object, 'is_approved', False):
                if self.request.POST.get('is_approval_request') == 'true':
                    self.object.status = 'APPROVAL_PENDING'
                    approver_id = self.request.POST.get('designated_approver')
                    if approver_id:
                        from users.models import User
                        try:
                            self.object.designated_approver = User.objects.get(pk=approver_id)
                        except User.DoesNotExist:
                            pass
                else:
                    form.add_error(None, ValidationError(f"Invoice cannot be saved because customer is {self.object.customer.customer_status}."))
                    return super().form_invalid(form)
            
            self.object.save()

            if self.object.status == 'DRAFT' and self.object.pk:
                delivery_notes = self.object.delivery_notes.all()
                if delivery_notes.exists():
                    addr_parts = [
                        self.object.snap_delivery_line1,
                        self.object.snap_delivery_line2,
                        self.object.snap_delivery_city,
                        self.object.snap_delivery_province,
                        self.object.snap_delivery_zip,
                    ]
                    if not any(addr_parts):
                        addr_parts = [
                            self.object.customer.delivery_address_line1,
                            self.object.customer.delivery_address_line2,
                            self.object.customer.delivery_city,
                            self.object.customer.delivery_province,
                            self.object.customer.delivery_zip_code,
                        ]
                    new_delivery_address = ", ".join([p for p in addr_parts if p])
                    delivery_notes.update(delivery_address=new_delivery_address)

            if items.is_valid():
                items.instance = self.object
                saved_items = items.save(commit=False)
                
                total = 0
                tax = 0
                tot_discount = 0
                for item in saved_items:
                    item.invoice = self.object
                    discount_amt = item.get_discount_amount
                    if self.object.customer.vat_enabled:
                        item.tax_amount = ((item.quantity * item.unit_price) - discount_amt) * Decimal('0.18')
                    else:
                        item.tax_amount = Decimal('0.00')
                        
                    item.line_total = (item.quantity * item.unit_price) - discount_amt + item.tax_amount
                    item.save()
                    
                for obj in items.deleted_objects:
                    obj.delete()
                    
                # Calculate aggregated values for invoice
                from .models import InvoiceItem
                current_items = InvoiceItem.objects.filter(invoice=self.object)
                gross_total = sum((item.quantity * item.unit_price) for item in current_items)
                line_discount = sum(item.get_discount_amount for item in current_items)
                subtotal = gross_total - line_discount
                
                custom_val = self.object.custom_discount_value or Decimal('0.00')
                if self.object.custom_discount_type == 'PERCENT':
                    global_discount = subtotal * (custom_val / Decimal('100.0'))
                else:
                    global_discount = custom_val
                    
                tot_discount = line_discount + global_discount
                subtotal -= global_discount
                if subtotal < Decimal('0.00'):
                    subtotal = Decimal('0.00')
                    
                if self.object.customer.vat_enabled:
                    tax = subtotal * Decimal('0.18')
                else:
                    tax = Decimal('0.00')
                
                total = subtotal + tax
                
                self.object.tax_amount = tax
                self.object.subtotal_amount = subtotal
                self.object.total_discount = tot_discount
                self.object.total_amount = total.quantize(Decimal('1.'), rounding=ROUND_HALF_UP)
                self.object.save()
                
                update_stock_reserves(self.object)
                
                if getattr(self.object, 'status', None) == 'APPROVAL_PENDING':
                    send_invoice_approval_email(self.object, self.request)
            else:
                return super().form_invalid(form)
            
        return super().form_valid(form)

class QuotationExportView(LoginRequiredMixin, ERPPermissionRequiredMixin, View):
    permission_required = 'sales.view_quotation'
    
    def get(self, request, *args, **kwargs):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="quotations.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Quotation Number', 'Customer', 'Creation Date', 'Salesperson', 'Valid Until', 'Total Amount', 'Status'])
        
        quotations = Quotation.objects.all().order_by('-creation_date')
        q = request.GET.get('q')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        salesperson_id = request.GET.get('salesperson')
        status = request.GET.get('status')
        
        from django.db.models import Q
        if q:
            quotations = quotations.filter(
                Q(quotation_number__icontains=q) |
                Q(customer__customer_name__icontains=q)
            )
        if date_from:
            quotations = quotations.filter(creation_date__date__gte=date_from)
        if date_to:
            quotations = quotations.filter(creation_date__date__lte=date_to)
        if salesperson_id:
            quotations = quotations.filter(salesperson_id=salesperson_id)
        if status:
            quotations = quotations.filter(status=status)
            
        for q_obj in quotations.order_by('-creation_date'):
            writer.writerow([q_obj.quotation_number, q_obj.customer.customer_name, q_obj.creation_date, q_obj.salesperson.username.title() if q_obj.salesperson else 'N/A', q_obj.valid_until, q_obj.total_amount, q_obj.get_status_display()])
            
        return response

class InvoiceExportView(LoginRequiredMixin, ERPPermissionRequiredMixin, View):
    permission_required = 'sales.view_invoice'

    COL_LABELS = {
        'invoice_number': 'Invoice Number',
        'invoice_date':   'Invoice Date',
        'due_date':       'Due Date',
        'invoice_type':   'Type',
        'customer':       'Customer',
        'salesperson':    'Salesperson',
        'status':         'Status',
        'delivery_date':  'Delivery Date',
        'subtotal':       'Sub Total (Ex-VAT)',
        'tax_amount':     'VAT Amount',
        'gross_amount':   'Gross Amount',
        'credit_note':    'Credit Note Value',
        'net_amount':     'Net Amount',
    }
    ALL_COLS = list(COL_LABELS.keys())

    def get(self, request, *args, **kwargs):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="invoices.csv"'

        # Which columns to include (default: all)
        requested_cols = request.GET.getlist('cols')
        cols = [c for c in self.ALL_COLS if c in requested_cols] if requested_cols else self.ALL_COLS

        writer = csv.writer(response)
        writer.writerow([self.COL_LABELS[c] for c in cols])

        invoices = Invoice.objects.all().order_by('-creation_date')
        q             = request.GET.get('q')
        date_from     = request.GET.get('date_from')
        date_to       = request.GET.get('date_to')
        salesperson_id = request.GET.get('salesperson')
        status        = request.GET.get('status')
        is_returned   = request.GET.get('is_returned')
        dashboard_mode = request.GET.get('dashboard_mode')

        from django.db.models import Q, Sum, F
        if q:
            invoices = invoices.filter(
                Q(invoice_number__icontains=q) |
                Q(customer__customer_name__icontains=q)
            )
        if date_from:
            invoices = invoices.filter(creation_date__date__gte=date_from)
        if date_to:
            invoices = invoices.filter(creation_date__date__lte=date_to)
        if salesperson_id:
            invoices = invoices.filter(salesperson_id=salesperson_id)
        if status:
            invoices = invoices.filter(status=status)
        if is_returned == 'true':
            invoices = invoices.filter(status__in=['CANCELLED', 'CANCEL_PENDING'], cancellation_reason__icontains='Customer Return')
        
        # Dashboard Match Mode filters out drafts and cancelled invoices
        if dashboard_mode == 'true':
            invoices = invoices.filter(status__in=[Invoice.Status.ISSUED, Invoice.Status.PAID])

        invoices = invoices.annotate(total_cn_value=Sum('credit_notes__items__credit_amount'))

        for inv in invoices:
            cn_value   = inv.total_cn_value or 0
            net_amount = inv.total_amount - cn_value
            # For accurate Ex-VAT Sub Total
            sub_total = inv.total_amount - inv.tax_amount

            row_map = {
                'invoice_number': inv.invoice_number,
                'invoice_date':   inv.creation_date.strftime('%Y-%m-%d') if inv.creation_date else '',
                'due_date':       inv.due_date.strftime('%Y-%m-%d') if inv.due_date else '',
                'invoice_type':   inv.get_invoice_type_display(),
                'customer':       inv.customer.customer_name,
                'salesperson':    inv.salesperson.get_full_name() or inv.salesperson.username if inv.salesperson else 'N/A',
                'status':         inv.get_status_display(),
                'delivery_date':  inv.delivery_date or '',
                'subtotal':       sub_total,
                'tax_amount':     inv.tax_amount,
                'gross_amount':   inv.total_amount,
                'credit_note':    cn_value,
                'net_amount':     net_amount,
            }
            writer.writerow([row_map[c] for c in cols])

        return response

import math

class InvoicePrintView(LoginRequiredMixin, ERPPermissionRequiredMixin, DetailView):
    model = Invoice
    template_name = 'sales/invoice_print.html'
    context_object_name = 'invoice'
    permission_required = 'sales.view_invoice'

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        if self.object.status == 'APPROVAL_PENDING':
            messages.error(request, "Cannot print an invoice that is pending approval.")
            return redirect('invoice_list')
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Total Value of Supply = total_amount - tax_amount 
        subtotal = self.object.total_amount - self.object.tax_amount
        context['total_value_supply'] = subtotal
        context['tax_amount'] = self.object.tax_amount
        context['total_amount'] = self.object.total_amount
        
        from decimal import Decimal, ROUND_HALF_UP
        rounded_total = self.object.total_amount.quantize(Decimal('1.'), rounding=ROUND_HALF_UP)
        context['rounded_total'] = rounded_total
        
        try:
            context['amount_in_words'] = num2words(int(rounded_total), lang='en').title() + " Rupees Only"
        except:
            context['amount_in_words'] = ""
        return context

class QuotationPrintView(LoginRequiredMixin, ERPPermissionRequiredMixin, DetailView):
    model = Quotation
    template_name = 'sales/quotation_print.html'
    context_object_name = 'quotation'
    permission_required = 'sales.view_quotation'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Total Value of Supply = total_amount - tax_amount
        subtotal = self.object.total_amount - self.object.tax_amount
        context['total_value_supply'] = subtotal
        context['tax_amount'] = self.object.tax_amount
        context['total_amount'] = self.object.total_amount
        
        from decimal import Decimal, ROUND_HALF_UP
        rounded_total = self.object.total_amount.quantize(Decimal('1.'), rounding=ROUND_HALF_UP)
        context['rounded_total'] = rounded_total
        
        try:
            context['amount_in_words'] = num2words(int(rounded_total), lang='en').title() + " Rupees Only"
        except:
            context['amount_in_words'] = ""
        return context

@login_required
def confirm_invoice_view(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    if request.method == 'POST':
        try:
            issue_invoice(invoice, request.user)
            messages.success(request, f"Invoice {invoice.invoice_number} issued. Stock deducted.")
        except Exception as e:
            messages.error(request, f"Error: {str(e)}")
    return redirect('invoice_list')

@login_required
def cancel_invoice_view(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    if request.method == 'POST':
        if invoice.status != 'ISSUED':
            messages.error(request, "Only Issued invoices can be cancelled.")
            return redirect('invoice_list')
            
        reason_type = request.POST.get('cancellation_reason_type', '')
        reason_text = request.POST.get('cancellation_reason', '')
        
        if reason_type == 'Other':
            reason = reason_text
        elif reason_type:
            reason = f"{reason_type}: {reason_text}" if reason_text else reason_type
        else:
            reason = reason_text
            
        approver_id = request.POST.get('designated_approver')
        if not reason or not approver_id:
            messages.error(request, "Reason and Approver are required.")
            return redirect('invoice_list')
            
        from users.models import User
        try:
            approver = User.objects.get(pk=approver_id)
        except User.DoesNotExist:
            messages.error(request, "Invalid approver selected.")
            return redirect('invoice_list')

        old_status = invoice.get_status_display()
        invoice.status = 'CANCEL_PENDING'
        invoice.cancellation_reason = reason
        invoice.designated_approver = approver
        invoice.save(update_fields=['status', 'cancellation_reason', 'designated_approver'])
        
        log_sales_event(
            obj=invoice,
            user=request.user,
            action="Cancellation Requested",
            old_value=old_status,
            new_value="Cancellation Pending",
            notes=f"Requested by {request.user.get_full_name()}. Reason: {reason}. Assigned to: {approver.get_full_name()}"
        )
        
        # Notify the specific approver
        from users.models import Notification
        Notification.objects.create(
            recipient=approver,
            notification_type='approval_required',
            title="Cancellation Approval Required",
            message=f"Cancellation requested for Invoice {invoice.invoice_number} by {request.user.get_full_name()}.",
            link=reverse('invoice_list'),
            action_approve_url=reverse('invoice_approve', kwargs={'pk': invoice.pk}),
            action_reject_url=reverse('invoice_reject', kwargs={'pk': invoice.pk})
        )
            
        messages.success(request, f"Cancellation request for {invoice.invoice_number} has been sent to {approver.get_full_name()} for approval.")
    return redirect('invoice_list')

@login_required
def request_edit_invoice_view(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    if request.method == 'POST':
        if invoice.status != 'ISSUED':
            messages.error(request, "Only Issued invoices can be edited.")
            return redirect('invoice_list')
            
        reason = request.POST.get('cancellation_reason')
        approver_id = request.POST.get('designated_approver')
        if not reason or not approver_id:
            messages.error(request, "Reason and Approver are required.")
            return redirect('invoice_list')
            
        from users.models import User
        try:
            approver = User.objects.get(pk=approver_id)
        except User.DoesNotExist:
            messages.error(request, "Invalid approver selected.")
            return redirect('invoice_list')

        old_status = invoice.get_status_display()
        invoice.status = 'EDIT_PENDING'
        invoice.cancellation_reason = reason
        invoice.designated_approver = approver
        invoice.save(update_fields=['status', 'cancellation_reason', 'designated_approver'])
        
        log_sales_event(
            obj=invoice,
            user=request.user,
            action="Edit Requested",
            old_value=old_status,
            new_value="Edit Pending",
            notes=f"Requested by {request.user.get_full_name()}. Reason: {reason}. Assigned to: {approver.get_full_name()}"
        )
        
        from users.models import Notification
        from django.core.mail import send_mail
        from django.conf import settings
        
        Notification.objects.create(
            recipient=approver,
            notification_type='approval_required',
            title="Edit Approval Required",
            message=f"Edit requested for Invoice {invoice.invoice_number} by {request.user.get_full_name()}.",
            link=reverse('invoice_list'),
            action_approve_url=reverse('invoice_approve', kwargs={'pk': invoice.pk}),
            action_reject_url=reverse('invoice_reject', kwargs={'pk': invoice.pk})
        )
        
        if approver.receive_email_alerts and approver.email:
            try:
                send_mail(
                    subject=f"Approval Required: Edit Invoice {invoice.invoice_number}",
                    message=f"{request.user.get_full_name()} has requested to edit Invoice {invoice.invoice_number}.\n\nReason: {reason}\n\nPlease log in to Everbolt ERP Action Center to approve or reject this request.",
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[approver.email],
                    fail_silently=True,
                )
            except Exception:
                pass
            
        messages.success(request, f"Edit request for {invoice.invoice_number} has been sent to {approver.get_full_name()} for approval.")
    return redirect('invoice_list')

@login_required
def approve_invoice_view(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    if not request.user.has_perm('sales.approve_invoice'):
        messages.error(request, "You do not have permission to approve invoices.")
        return redirect('invoice_list')
        
    if request.method == 'POST':
        reviewer_notes = request.POST.get('reviewer_notes', '')
        
        # Mark associated notifications as read
        from users.models import Notification
        Notification.objects.filter(
            recipient=request.user, 
            notification_type='approval_required',
            message__contains=f"Invoice {invoice.invoice_number}"
        ).update(is_read=True)
        
        if invoice.status == 'APPROVAL_PENDING':
            old_status = invoice.get_status_display()
            invoice.status = 'DRAFT'
            invoice.is_approved = True
            invoice.reviewer_notes = reviewer_notes
            invoice.save(update_fields=['status', 'is_approved', 'reviewer_notes'])
            
            log_sales_event(
                obj=invoice,
                user=request.user,
                action="Invoice Approved",
                old_value=old_status,
                new_value=invoice.get_status_display(),
                notes=f"Manager Notes: {reviewer_notes}"
            )
            
            # Notify creator
            from users.models import Notification
            if invoice.salesperson:
                msg = f"Your invoice {invoice.invoice_number} has been approved."
                if reviewer_notes:
                    msg += f" Manager Notes: {reviewer_notes}"
                
                Notification.objects.create(
                    recipient=invoice.salesperson,
                    title="Invoice Approved",
                    message=msg,
                    link=reverse('invoice_edit', kwargs={'pk': invoice.pk})
                )
            
            messages.success(request, f"Invoice {invoice.invoice_number} has been approved and moved to Draft.")
            
        elif invoice.status == 'CANCEL_PENDING':
            # Check for permissions for cancellation approval
            if not request.user.has_perm('sales.approve_invoice'):
                messages.error(request, "You do not have permission to approve cancellations.")
                return redirect('invoice_list')
                
            old_status = invoice.get_status_display()
            try:
                from .services import cancel_invoice as service_cancel_invoice
                service_cancel_invoice(invoice, request.user)
                invoice.reviewer_notes = reviewer_notes
                invoice.save(update_fields=['reviewer_notes'])
                
                log_sales_event(
                    obj=invoice,
                    user=request.user,
                    action="Cancellation Approved",
                    old_value=old_status,
                    new_value="Cancelled",
                    notes=f"Manager Notes: {reviewer_notes}"
                )
                
                # Notify creator
                from users.models import Notification
                if invoice.salesperson:
                    msg = f"Invoice {invoice.invoice_number} cancellation has been approved. Stock restored."
                    if reviewer_notes:
                        msg += f" Manager Notes: {reviewer_notes}"
                        
                    Notification.objects.create(
                        recipient=invoice.salesperson,
                        title="Cancellation Approved",
                        message=msg,
                        link=reverse('invoice_list')
                    )
                
                messages.success(request, f"Cancellation for {invoice.invoice_number} has been approved. Stock has been restored.")
            except Exception as e:
                messages.error(request, f"Error during cancellation: {str(e)}")
                
        elif invoice.status == 'EDIT_PENDING':
            # Check for permissions for edit approval
            if not request.user.has_perm('sales.approve_invoice'):
                messages.error(request, "You do not have permission to approve edits.")
                return redirect('invoice_list')
                
            old_status = invoice.get_status_display()
            try:
                from .services import restore_stock
                # Use restore_stock to update inventory without setting status to CANCELLED
                restore_stock(invoice, request.user, "Edit Approved (Stock Restored)")
                
                invoice.status = 'DRAFT'
                invoice.reviewer_notes = reviewer_notes
                invoice.cancellation_reason = ''  # Clear reason
                invoice.save(update_fields=['status', 'reviewer_notes', 'cancellation_reason'])
                
                log_sales_event(
                    obj=invoice,
                    user=request.user,
                    action="Edit Approved",
                    old_value=old_status,
                    new_value="Draft",
                    notes=f"Stock restored. Manager Notes: {reviewer_notes}"
                )
                
                # Notify creator
                from users.models import Notification
                if invoice.salesperson:
                    msg = f"Invoice {invoice.invoice_number} edit has been approved. It is now a Draft and stock is restored."
                    if reviewer_notes:
                        msg += f" Manager Notes: {reviewer_notes}"

                    Notification.objects.create(
                        recipient=invoice.salesperson,
                        title="Edit Approved",
                        message=msg,
                        link=reverse('invoice_edit', kwargs={'pk': invoice.pk})
                    )
                
                messages.success(request, f"Edit for {invoice.invoice_number} has been approved. Stock restored and invoice is now a Draft.")
            except Exception as e:
                messages.error(request, f"Error during edit approval: {str(e)}")
        else:
            messages.warning(request, "This invoice is not pending any approval.")
    return redirect('invoice_list')

@login_required
def reject_invoice_view(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    if not request.user.has_perm('sales.approve_invoice'):
        messages.error(request, "You do not have permission to reject invoices.")
        return redirect('invoice_list')
        
    if request.method == 'POST':
        reviewer_notes = request.POST.get('reviewer_notes', '')
        
        # Mark associated notifications as read
        from users.models import Notification
        Notification.objects.filter(
            recipient=request.user, 
            notification_type='approval_required',
            message__contains=f"Invoice {invoice.invoice_number}"
        ).update(is_read=True)
        
        if invoice.status == 'APPROVAL_PENDING':
            old_status = invoice.get_status_display()
            invoice.status = 'DRAFT' # Correct: Back to Draft for salesperson to fix or issue
            invoice.is_approved = False # Still not approved if rejected
            invoice.reviewer_notes = reviewer_notes
            invoice.save(update_fields=['status', 'is_approved', 'reviewer_notes'])
            
            log_sales_event(
                obj=invoice,
                user=request.user,
                action="Invoice Rejected",
                old_value=old_status,
                new_value=invoice.get_status_display(),
                notes=f"Manager Notes: {reviewer_notes}"
            )
            
            # Notify creator
            from users.models import Notification
            if invoice.salesperson:
                msg = f"Your invoice {invoice.invoice_number} has been rejected."
                if reviewer_notes:
                    msg += f" Manager Notes: {reviewer_notes}"
                else:
                    msg += " No reason provided."

                Notification.objects.create(
                    recipient=invoice.salesperson,
                    title="Invoice Rejected",
                    message=msg,
                    link=reverse('invoice_edit', kwargs={'pk': invoice.pk})
                )
            
            messages.warning(request, f"Invoice {invoice.invoice_number} has been rejected and moved back to Draft.")
            
        elif invoice.status == 'CANCEL_PENDING':
            # Check for permissions for cancellation rejection
            if not request.user.has_perm('sales.approve_invoice'):
                messages.error(request, "You do not have permission to reject cancellations.")
                return redirect('invoice_list')
                
            old_status = invoice.get_status_display()
            invoice.status = 'ISSUED' # Return to Issued
            invoice.reviewer_notes = reviewer_notes
            invoice.save(update_fields=['status', 'reviewer_notes'])
            
            log_sales_event(
                obj=invoice,
                user=request.user,
                action="Cancellation Rejected",
                old_value=old_status,
                new_value="Issued",
                notes=f"Manager Notes: {reviewer_notes}"
            )
            
            # Notify creator
            from users.models import Notification
            if invoice.salesperson:
                msg = f"Cancellation request for Invoice {invoice.invoice_number} was rejected. Status returned to Issued."
                if reviewer_notes:
                    msg += f" Manager Notes: {reviewer_notes}"

                Notification.objects.create(
                    recipient=invoice.salesperson,
                    title="Cancellation Rejected",
                    message=msg,
                    link=reverse('invoice_list')
                )
                
            messages.warning(request, f"Cancellation request for {invoice.invoice_number} has been rejected.")
            
        elif invoice.status == 'EDIT_PENDING':
            # Check for permissions for edit rejection
            if not request.user.has_perm('sales.approve_invoice'):
                messages.error(request, "You do not have permission to reject edits.")
                return redirect('invoice_list')
                
            old_status = invoice.get_status_display()
            invoice.status = 'ISSUED' # Return to Issued
            invoice.reviewer_notes = reviewer_notes
            invoice.cancellation_reason = ''
            invoice.save(update_fields=['status', 'reviewer_notes', 'cancellation_reason'])
            
            log_sales_event(
                obj=invoice,
                user=request.user,
                action="Edit Rejected",
                old_value=old_status,
                new_value="Issued",
                notes=f"Manager Notes: {reviewer_notes}"
            )
            
            from users.models import Notification
            if invoice.salesperson:
                msg = f"Edit request for Invoice {invoice.invoice_number} was rejected. The invoice remains in Issued status."
                if reviewer_notes:
                    msg += f" Manager Reason: {reviewer_notes}"

                Notification.objects.create(
                    recipient=invoice.salesperson,
                    title="Edit Request Denied",
                    message=msg,
                    link=reverse('invoice_list')
                )
                
            messages.warning(request, f"Edit request for {invoice.invoice_number} has been rejected.")
        else:
            messages.warning(request, "This invoice is not pending any approval.")
            
    return redirect('invoice_list')

@login_required
def quotation_mark_sent_view(request, pk):
    """Mark a DRAFT quotation as SENT (i.e. delivered to customer)."""
    quotation = get_object_or_404(Quotation, pk=pk)
    if request.method == 'POST':
        if quotation.status == 'DRAFT':
            old_status = quotation.get_status_display()
            quotation.status = 'SENT'
            quotation.save(update_fields=['status'])
            
            log_sales_event(
                obj=quotation,
                user=request.user,
                action="Quotation Sent",
                old_value=old_status,
                new_value=quotation.get_status_display()
            )
            
            messages.success(request, f"Quotation {quotation.quotation_number} marked as Sent.")
        else:
            messages.warning(request, "Only DRAFT quotations can be marked as Sent.")
    return redirect('quotation_list')

@login_required
def quotation_cancel_view(request, pk):
    """Cancel a DRAFT or SENT quotation."""
    quotation = get_object_or_404(Quotation, pk=pk)
    if request.method == 'POST':
        if quotation.status in ['DRAFT', 'SENT']:
            old_status = quotation.get_status_display()
            quotation.status = 'CANCELLED'
            quotation.save(update_fields=['status'])
            
            log_sales_event(
                obj=quotation,
                user=request.user,
                action="Quotation Cancelled",
                old_value=old_status,
                new_value=quotation.get_status_display()
            )
            
            messages.success(request, f"Quotation {quotation.quotation_number} cancelled.")
        else:
            messages.warning(request, "Only DRAFT or SENT quotations can be cancelled.")
    return redirect('quotation_list')


@login_required
@permission_required('sales.add_invoice', raise_exception=True)
def convert_quotation_view(request, pk):
    """Converts a Quotation into a Draft Invoice."""
    quotation = get_object_or_404(Quotation, pk=pk)
    
    if quotation.is_converted:
        messages.warning(request, "This quotation has already been converted to an invoice.")
        return redirect('quotation_list')
        
    custom_items = quotation.items.filter(product__isnull=True)
    if custom_items.exists():
        item_names = ", ".join([item.description for item in custom_items])
        messages.error(request, f"This quotation contains custom items ({item_names}). You must edit the quotation and link official products from the inventory before converting it to an invoice.")
        return redirect('quotation_list')
        
    with transaction.atomic():
        # Create Invoice Header
        invoice = Invoice.objects.create(
            customer=quotation.customer,
            salesperson=quotation.customer.assigned_sales_officer or request.user,
            total_amount=quotation.total_amount,
            subtotal_amount=quotation.subtotal_amount,
            tax_amount=quotation.tax_amount,
            total_discount=quotation.total_discount,
            custom_discount_type=quotation.custom_discount_type,
            custom_discount_value=quotation.custom_discount_value,
            notes=f"Converted from Quotation {quotation.quotation_number}. " + (quotation.notes or ""),
            status='DRAFT'
        )
        
        # Create Invoice Items
        from .models import InvoiceItem
        for q_item in quotation.items.all():
            InvoiceItem.objects.create(
                invoice=invoice,
                product=q_item.product,
                quantity=q_item.quantity,
                unit_price=q_item.unit_price,
                discount_type=q_item.discount_type,
                discount=q_item.discount,
                tax_amount=q_item.tax_amount,
                line_total=q_item.line_total
            )
            
        # Update Quotation status
        quotation.status = 'CONVERTED'
        quotation.is_converted = True
        quotation.save(update_fields=['status', 'is_converted'])
        
        log_sales_event(
            obj=quotation,
            user=request.user,
            action="Converted to Invoice",
            new_value=invoice.invoice_number
        )
        
        log_sales_event(
            obj=invoice,
            user=request.user,
            action="Created from Quotation",
            old_value=quotation.quotation_number
        )
        
        messages.success(request, f"Quotation {quotation.quotation_number} converted to Invoice {invoice.invoice_number} successfully.")
        return redirect('invoice_edit', pk=invoice.pk)

from django.http import JsonResponse

@login_required
def customer_search_ajax(request):
    """API endpoint for Select2 AJAX customer search."""
    q = request.GET.get('q', '')
    customers = Customer.objects.filter(
        Q(customer_name__icontains=q) | 
        Q(company_name__icontains=q) |
        Q(phone_number__icontains=q)
    )[:20]
    
    results = [
        {'id': c.id, 'text': f"{c.customer_name} ({c.company_name or 'No Company'})"} 
        for c in customers
    ]
    return JsonResponse({'results': results})

@login_required
def product_search_ajax(request):
    """API endpoint for Select2 AJAX product search."""
    q = request.GET.get('q', '')
    products = Product.objects.filter(
        Q(name__icontains=q) | 
        Q(product_id__icontains=q)
    )[:20]
    results = []
    for p in products:
        tiers_data = []
        tiers_qs = p.price_tiers.order_by('min_quantity')
        
        if tiers_qs.exists():
            tier_list = list(tiers_qs)
            # Base fallback tier if lowest tier starts > 1
            if tier_list[0].min_quantity > 1:
                tiers_data.append({
                    'min_quantity': 1,
                    'price': float(p.selling_price)
                })
            
            for t in tier_list:
                tiers_data.append({
                    'min_quantity': t.min_quantity,
                    'price': float(t.price)
                })
                
        # Sort descending for the frontend logic
        tiers_data.sort(key=lambda x: x['min_quantity'], reverse=True)
            
        results.append({
            'id': p.id, 
            'text': f"[{p.product_id}] {p.name}",
            'price': float(p.selling_price),
            'stock': float(p.available_stock),
            'price_tiers': tiers_data
        })
        
    return JsonResponse({'results': results})
class DeliveryNoteListView(LoginRequiredMixin, ERPPermissionRequiredMixin, ListView):
    model = DeliveryNote
    template_name = 'sales/delivery_note_list.html'
    context_object_name = 'delivery_notes'
    paginate_by = 20
    permission_required = 'sales.view_deliverynote'
    
    def get_queryset(self):
        qs = super().get_queryset().select_related('invoice__customer', 'delivered_by').order_by('-created_at')
        
        # Status Filter
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        
        # Delivered By Filter
        delivered_by = self.request.GET.get('delivered_by')
        if delivered_by:
            qs = qs.filter(delivered_by_id=delivered_by)
            
        # Date Range Filter
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        # Delivery Date Range Filter
        delivery_date_from = self.request.GET.get('delivery_date_from')
        delivery_date_to = self.request.GET.get('delivery_date_to')
        if delivery_date_from:
            qs = qs.filter(delivery_date__gte=delivery_date_from)
        if delivery_date_to:
            qs = qs.filter(delivery_date__lte=delivery_date_to)

        # Unified Search (DN Details, Invoice, Customer)
        q = self.request.GET.get('q')
        if q:
            from django.db.models import Q
            qs = qs.filter(
                Q(dn_number__icontains=q) |
                Q(invoice__invoice_number__icontains=q) |
                Q(customer_name__icontains=q)
            )
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['model_name'] = 'DeliveryNote'
        
        from users.models import User
        context['delivery_officers'] = User.objects.filter(is_active=True, is_delivery_officer=True)
        
        if self.request.user.is_authenticated:
            context['saved_filters'] = SavedFilter.objects.filter(
                user=self.request.user, 
                model_name='DeliveryNote'
            )
        else:
            context['saved_filters'] = []
        return context

class DeliveryNoteDetailView(LoginRequiredMixin, ERPPermissionRequiredMixin, DetailView):
    model = DeliveryNote
    template_name = 'sales/delivery_note_detail.html'
    context_object_name = 'dn'
    permission_required = 'sales.view_deliverynote'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ct = ContentType.objects.get_for_model(DeliveryNote)
        context['dn_audit_history'] = SalesAuditLog.objects.filter(
            content_type=ct, 
            object_id=self.object.id
        ).order_by('-timestamp')
        return context

class DeliveryNoteCreateView(LoginRequiredMixin, ERPPermissionRequiredMixin, CreateView):
    model = DeliveryNote
    form_class = DeliveryNoteForm
    template_name = 'sales/delivery_note_form.html'
    permission_required = 'sales.add_deliverynote'

    def get_success_url(self):
        return reverse('delivery_note_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        with transaction.atomic():
            self.object = form.save(commit=False)
            invoice = self.object.invoice

            # Fix #8: Only allow DN creation against ISSUED invoices
            if invoice.status != 'ISSUED':
                form.add_error(
                    'invoice',
                    f"Delivery Notes can only be created for ISSUED invoices. "
                    f"'{invoice.invoice_number}' is currently {invoice.get_status_display()}."
                )
                return self.form_invalid(form)

            # Fallback for missing fields if JS failed or customer lacks them
            if not self.object.customer_name:
                self.object.customer_name = invoice.customer.company_name or invoice.customer.customer_name
            if not self.object.delivery_address:
                if invoice.snap_delivery_line1 or invoice.snap_delivery_city:
                    addr_parts = [
                        invoice.snap_delivery_line1,
                        invoice.snap_delivery_line2,
                        invoice.snap_delivery_city,
                        invoice.snap_delivery_province,
                        invoice.snap_delivery_zip,
                    ]
                else:
                    addr_parts = [
                        invoice.customer.delivery_address_line1,
                        invoice.customer.delivery_address_line2,
                        invoice.customer.delivery_city,
                        invoice.customer.delivery_province,
                        invoice.customer.delivery_zip_code,
                    ]
                self.object.delivery_address = ", ".join([p for p in addr_parts if p])
            if not self.object.delivery_date:
                from django.utils import timezone
                self.object.delivery_date = invoice.delivery_date or timezone.now().date()

            self.object.save()

            # Fix #9: Copy items with invoiced_quantity stored as a cap.
            # Delivery quantity is set to the full invoiced amount by default
            # (user can reduce it for partial delivery, but never exceed it).
            for item in invoice.items.all():
                DeliveryNoteItem.objects.create(
                    delivery_note=self.object,
                    product=item.product,
                    quantity=item.quantity,           # delivered qty (default = full)
                    invoiced_quantity=item.quantity,  # hard cap — never exceed this
                )

            # Deduct stock instantly upon DN creation
            deduct_dn_stock(self.object, self.request.user)

            messages.success(self.request, f"Delivery Note {self.object.dn_number} created successfully. Stock has been deducted.")

            log_sales_event(
                obj=self.object,
                user=self.request.user,
                action="Delivery Note Created",
                new_value=self.object.get_status_display(),
                notes=f"Linked to Invoice {invoice.invoice_number}"
            )

            return super().form_valid(form)

class DeliveryNoteUpdateView(LoginRequiredMixin, ERPPermissionRequiredMixin, UpdateView):
    model = DeliveryNote
    form_class = DeliveryNoteForm
    template_name = 'sales/delivery_note_form.html'
    permission_required = 'sales.change_deliverynote'

    def get_success_url(self):
        return reverse('delivery_note_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        with transaction.atomic():
            old_delivered_by = self.get_object().delivered_by
            self.object = form.save()
            
            if old_delivered_by != self.object.delivered_by:
                log_sales_event(
                    obj=self.object,
                    user=self.request.user,
                    action="Delivery Person Updated",
                    old_value=old_delivered_by.get_full_name() if old_delivered_by else "None",
                    new_value=self.object.delivered_by.get_full_name() if self.object.delivered_by else "None",
                    notes="Updated Delivery Officer"
                )

            messages.success(self.request, f"Delivery Note {self.object.dn_number} updated successfully.")
            return super().form_valid(form)

class DeliveryNoteDeleteView(LoginRequiredMixin, AdminRequiredMixin, DeleteView):
    model = DeliveryNote
    template_name = 'sales/delivery_note_confirm_delete.html'
    success_url = reverse_lazy('delivery_note_list')
    permission_required = 'sales.delete_deliverynote'

    def form_valid(self, form):
        from .services import restore_dn_stock
        with transaction.atomic():
            # Ensure the invoice stock_deducted flag is reverted and stock is added back
            restore_dn_stock(self.object, self.request.user, remark_prefix="Delivery Note Deleted")
            
            # Log deletion on the invoice
            log_sales_event(
                obj=self.object.invoice,
                user=self.request.user,
                action="Delivery Note Deleted",
                new_value="N/A",
                notes=f"Delivery Note {self.object.dn_number} deleted by Admin."
            )
            
            messages.success(self.request, f"Delivery Note {self.object.dn_number} successfully deleted and stock reversed.")
            return super().form_valid(form)
@login_required
def get_invoice_details(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    customer = invoice.customer

    # Use the invoice's own snapshotted delivery address if available.
    # Fall back to the customer's current address for backward compatibility.
    if invoice.snap_delivery_line1 or invoice.snap_delivery_city:
        addr_parts = [
            invoice.snap_delivery_line1,
            invoice.snap_delivery_line2,
            invoice.snap_delivery_city,
            invoice.snap_delivery_province,
            invoice.snap_delivery_zip,
        ]
    else:
        addr_parts = [
            customer.delivery_address_line1,
            customer.delivery_address_line2,
            customer.delivery_city,
            customer.delivery_province,
            customer.delivery_zip_code,
        ]
    address = ", ".join([p for p in addr_parts if p])

    items = []
    for item in invoice.items.all():
        items.append({
            'product_name': item.product.name,
            'quantity': str(item.quantity),
            'product_id': item.product.product_id
        })

    data = {
        'customer_name': customer.company_name or customer.customer_name,
        'delivery_address': address,
        'delivery_date': invoice.delivery_date.isoformat() if invoice.delivery_date else '',
        'items': items
    }
    return JsonResponse(data)

@login_required
@permission_required('sales.change_dn_status', raise_exception=True)
def update_dn_status(request, pk):
    dn = get_object_or_404(DeliveryNote, pk=pk)
    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status in dict(DeliveryNote.Status.choices):
            old_display = dn.get_status_display()
            dn.status = new_status
            dn.save()
            
            if new_status == 'FAILED' and old_display != 'Failed':
                restore_dn_stock(dn, request.user)
                messages.warning(request, f"Stock for {dn.dn_number} has been restored to inventory.")
            elif old_display == 'Failed' and new_status != 'FAILED':
                # Re-deduct if moved out of FAILED
                deduct_dn_stock(dn, request.user)
                messages.success(request, f"Stock for {dn.dn_number} has been deducted again.")
            
            log_sales_event(
                obj=dn,
                user=request.user,
                action="Status Updated",
                old_value=old_display,
                new_value=dn.get_status_display(),
                notes=f"Manual status change from list/detail view."
            )
            
            messages.success(request, f"Status of {dn.dn_number} updated to {dn.get_status_display()}.")
    return redirect('delivery_note_list')


# ─────────────────────────────────────────────────────────────────────────────
# RETURNS & CREDIT NOTES
# ─────────────────────────────────────────────────────────────────────────────

class ReturnListView(LoginRequiredMixin, ERPPermissionRequiredMixin, ListView):
    model = Return
    template_name = 'sales/return_list.html'
    context_object_name = 'returns'
    paginate_by = 20
    permission_required = 'sales.view_return'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['model_name'] = 'Return'
        try:
            from users.models import SavedFilter
            context['saved_filters'] = SavedFilter.objects.filter(user=self.request.user, model_name='Return')
        except ImportError:
            context['saved_filters'] = []
        context['reasons'] = Return.ReturnReason.choices
        return context

    def get_queryset(self):
        qs = Return.objects.select_related('original_invoice', 'created_by').prefetch_related('items', 'items__product').order_by('-created_date')
        q = self.request.GET.get('q')
        if q:
            from django.db.models import Q
            qs = qs.filter(
                Q(return_number__icontains=q) |
                Q(original_invoice__invoice_number__icontains=q) |
                Q(items__product__name__icontains=q)
            ).distinct()
        status = self.request.GET.get('status')
        if status == 'processed':
            qs = qs.filter(stock_updated=True)
        elif status == 'pending':
            qs = qs.filter(stock_updated=False)
            
        reason = self.request.GET.get('reason')
        if reason:
            qs = qs.filter(items__reason=reason).distinct()
            
        return qs


from .forms import ReturnForm, ReturnItemFormSet

@login_required
@permission_required('sales.add_return', raise_exception=True)
def return_create_view(request, invoice_pk):
    """Create a Return with multiple items against a specific ISSUED/PAID invoice."""
    invoice = get_object_or_404(Invoice, pk=invoice_pk)

    if invoice.status not in ['ISSUED', 'PAID']:
        messages.error(request, "Returns can only be raised against Issued or Paid invoices.")
        return redirect('invoice_list')

    # Calculate already returned quantities
    returned_quantities = {}
    for ret in invoice.returns.filter(stock_updated=True):
        for ri in ret.items.all():
            returned_quantities[ri.product_id] = returned_quantities.get(ri.product_id, 0) + ri.quantity

    if request.method == 'POST':
        form = ReturnForm(request.POST)
        formset = ReturnItemFormSet(request.POST)
        
        if form.is_valid() and formset.is_valid():
            # Check if any items were submitted
            has_items = False
            for inline_form in formset:
                if inline_form.cleaned_data and not inline_form.cleaned_data.get('DELETE', False):
                    has_items = True
                    break
            
            if not has_items:
                messages.error(request, "You must select at least one product to return.")
            else:
                # Validate that quantities don't exceed remaining invoiced quantity
                invalid_items = []
                for inline_form in formset:
                    if inline_form.cleaned_data and not inline_form.cleaned_data.get('DELETE', False):
                        product = inline_form.cleaned_data.get('product')
                        qty = inline_form.cleaned_data.get('quantity', 0)
                        if product and qty > 0:
                            already_returned = returned_quantities.get(product.id, 0)
                            inv_item = invoice.items.filter(product=product).first()
                            max_qty = inv_item.quantity - already_returned if inv_item else 0
                            if qty > max_qty:
                                invalid_items.append(f"{product.name} (Max remaining: {max_qty})")
                
                if invalid_items:
                    messages.error(request, f"Cannot return more than the remaining invoiced amount for: {', '.join(invalid_items)}")
                else:
                    try:
                        ret = form.save(commit=False)
                        ret.original_invoice = invoice
                        ret.created_by = request.user
                        ret.save()
                        
                        formset.instance = ret
                        formset.save()

                        # Immediately process: restore stock + generate credit note
                        credit_note = process_return(ret, request.user)
                        messages.success(
                            request,
                            f"Return {ret.return_number} processed. Stock restored. "
                            f"Credit Note {credit_note.credit_note_number} issued (Rs {credit_note.total_credit_amount})."
                        )
                        return redirect('credit_note_list')
                    except Exception as e:
                        messages.error(request, f"Error processing return: {e}")
                        # If error occurs, we should technically delete the Return we just saved, but atomic tx in process_return will handle partials. If it failed before process_return, we might have an orphaned Return.
                        # Best to delete it if process_return fails.
                        try:
                            ret.delete()
                        except:
                            pass
        else:
            messages.error(request, "Please correct the errors in the form.")
    else:
        form = ReturnForm()
        initial_data = []
        for item in invoice.items.select_related('product').all():
            already_returned = returned_quantities.get(item.product_id, 0)
            max_qty = item.quantity - already_returned
            if max_qty > 0:
                initial_data.append({
                    'product': item.product.id,
                    'quantity': max_qty,
                    'unit_price': item.unit_price,
                })
        
        from django.forms import inlineformset_factory
        from .models import Return, ReturnItem
        from .forms import ReturnItemForm
        ReturnItemFormSetPrefilled = inlineformset_factory(
            Return, ReturnItem, form=ReturnItemForm,
            extra=len(initial_data), can_delete=True
        )
        formset = ReturnItemFormSetPrefilled(initial=initial_data)

    # Pass invoice items to the template so JS can fetch unit prices easily
    items_data = []
    for item in invoice.items.select_related('product').all():
        already_returned = returned_quantities.get(item.product_id, 0)
        max_qty = item.quantity - already_returned
        
        if max_qty > 0:
            items_data.append({
                'product_id': item.product.id,
                'name': item.product.name,
                'code': item.product.product_id,
                'unit_price': float(item.unit_price),
                'max_quantity': max_qty
            })

    import json
    return render(request, 'sales/return_form.html', {
        'form': form,
        'formset': formset,
        'invoice': invoice,
        'items_data_json': json.dumps(items_data),
    })


class CreditNoteListView(LoginRequiredMixin, ERPPermissionRequiredMixin, ListView):
    model = CreditNote
    template_name = 'sales/credit_note_list.html'
    context_object_name = 'credit_notes'
    paginate_by = 20
    permission_required = 'sales.view_return'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['model_name'] = 'CreditNote'
        try:
            from users.models import SavedFilter
            context['saved_filters'] = SavedFilter.objects.filter(user=self.request.user, model_name='CreditNote')
        except ImportError:
            context['saved_filters'] = []
        return context

    def get_queryset(self):
        qs = CreditNote.objects.select_related('customer', 'original_invoice').prefetch_related('items', 'items__product').order_by('-issued_date')
        q = self.request.GET.get('q')
        if q:
            from django.db.models import Q
            qs = qs.filter(
                Q(credit_note_number__icontains=q) |
                Q(original_invoice__invoice_number__icontains=q) |
                Q(customer__customer_name__icontains=q) |
                Q(items__product__name__icontains=q)
            ).distinct()
            
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        if date_from:
            qs = qs.filter(issued_date__date__gte=date_from)
        if date_to:
            qs = qs.filter(issued_date__date__lte=date_to)
            
        return qs


from .forms import CreditNoteForm, CreditNoteItemFormSet

class CreditNoteUpdateView(LoginRequiredMixin, ERPPermissionRequiredMixin, UpdateView):
    model = CreditNote
    form_class = CreditNoteForm
    template_name = 'sales/credit_note_form.html'
    permission_required = 'sales.change_creditnote'
    
    def get_success_url(self):
        return reverse('credit_note_list')
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['formset'] = CreditNoteItemFormSet(self.request.POST, instance=self.object)
        else:
            context['formset'] = CreditNoteItemFormSet(instance=self.object)
        return context
        
    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                self.object = form.save()
                formset.instance = self.object
                formset.save()
                
                log_sales_event(
                    obj=self.object.return_record,
                    user=self.request.user,
                    action="Credit Note Updated",
                    new_value=str(self.object.total_credit_amount),
                    notes="Updated Credit Note details and amounts."
                )
                
                messages.success(self.request, f"Credit Note {self.object.credit_note_number} updated successfully.")
                return super().form_valid(form)
        else:
            return self.form_invalid(form)

@login_required
def credit_note_print_view(request, pk):
    """Printable Credit Note view."""
    from decimal import Decimal
    cn = get_object_or_404(CreditNote, pk=pk)
    
    total_credit = cn.total_credit_amount
    tax_amount = Decimal('0.00')
    if cn.original_invoice.tax_amount > 0:
        tax_amount = total_credit * Decimal('0.18')
        
    total_with_tax = total_credit + tax_amount
    
    context = {
        'cn': cn,
        'tax_amount': tax_amount,
        'total_with_tax': total_with_tax,
    }
    return render(request, 'sales/credit_note_print.html', context)


class DeliveryNotePrintView(LoginRequiredMixin, ERPPermissionRequiredMixin, DetailView):
    model = DeliveryNote
    template_name = 'sales/delivery_note_print.html'
    context_object_name = 'dn'
    permission_required = 'sales.view_deliverynote'

# =============================================================================
# Dashboard & Tool Views
# =============================================================================

class DeliveryDashboardView(LoginRequiredMixin, ERPPermissionRequiredMixin, TemplateView):
    template_name = 'sales/delivery_dashboard.html'
    permission_required = 'sales.view_deliverynote'

import json

class OrderGeneratorView(LoginRequiredMixin, ERPPermissionRequiredMixin, TemplateView):
    template_name = 'sales/tools/order_generator.html'
    permission_required = 'sales.add_quotation'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from inventory.models import Product
        from crm.models import Customer
        
        products = Product.objects.filter(status=True).order_by('category', 'tea_type', 'name')
        customers = Customer.objects.filter(customer_status='ACTIVE').order_by('customer_name')
        
        categories = []
        product_data = {}
        
        for p in products:
            cat = (p.tea_type if p.category == 'Tea' and p.tea_type else p.category).upper()
            if cat not in categories:
                categories.append(cat)
                product_data[cat] = {}
                
            tiers_qs = p.price_tiers.order_by('min_quantity')
            if tiers_qs.exists():
                tiers = []
                tier_list = list(tiers_qs)
                
                # If first tier starts > 1, add a base tier
                if tier_list[0].min_quantity > 1:
                    tiers.append({
                        "min": 1, 
                        "max": tier_list[0].min_quantity - 1, 
                        "price": float(p.selling_price)
                    })
                    
                for i, tier in enumerate(tier_list):
                    tier_dict = {"min": tier.min_quantity, "price": float(tier.price)}
                    if i + 1 < len(tier_list):
                        tier_dict["max"] = tier_list[i+1].min_quantity - 1
                    tiers.append(tier_dict)
                product_data[cat][p.name] = tiers
            else:
                product_data[cat][p.name] = float(p.selling_price)
                
        customer_details = {}
        for c in customers:
            customer_details[c.customer_name] = {
                'address': c.billing_address.replace('\n', ', '),
                'phone': c.phone
            }
                
        context['customers'] = customers
        context['customer_data_json'] = json.dumps(customer_details)
        context['category_order_json'] = json.dumps(categories)
        context['product_data_json'] = json.dumps(product_data)
        return context

class CourierCalculatorView(LoginRequiredMixin, ERPPermissionRequiredMixin, TemplateView):
    template_name = 'sales/tools/courier_calculator.html'
    permission_required = 'sales.view_deliverynote'

class DeliveryNoteExportView(LoginRequiredMixin, ERPPermissionRequiredMixin, View):
    permission_required = 'sales.view_deliverynote'
    
    def get(self, request, *args, **kwargs):
        import openpyxl
        from openpyxl.utils import get_column_letter

        qs = DeliveryNote.objects.all().order_by('-created_at')
        
        # Status Filter
        status = request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        
        # Delivered By Filter
        delivered_by = request.GET.get('delivered_by')
        if delivered_by:
            qs = qs.filter(delivered_by_id=delivered_by)
            
        # Date Range Filter
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        # Delivery Date Range Filter
        delivery_date_from = request.GET.get('delivery_date_from')
        delivery_date_to = request.GET.get('delivery_date_to')
        if delivery_date_from:
            qs = qs.filter(delivery_date__gte=delivery_date_from)
        if delivery_date_to:
            qs = qs.filter(delivery_date__lte=delivery_date_to)

        # Unified Search
        q = request.GET.get('q')
        if q:
            from django.db.models import Q
            qs = qs.filter(
                Q(dn_number__icontains=q) |
                Q(invoice__invoice_number__icontains=q) |
                Q(customer_name__icontains=q)
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Delivery Notes Export"
        
        headers = [
            'DN Number', 'Invoice Number', 'Customer', 'Delivery Address', 
            'Target Delivery Date', 'Delivered By', 'Status', 'Created At', 'Remarks'
        ]
        ws.append(headers)

        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
            
        for dn in qs:
            ws.append([
                dn.dn_number,
                dn.invoice.invoice_number,
                dn.customer_name,
                dn.delivery_address,
                dn.delivery_date.strftime('%Y-%m-%d') if dn.delivery_date else '',
                dn.delivery_officer_name,
                dn.get_status_display(),
                dn.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                dn.remarks or ''
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="delivery_notes.xlsx"'
        wb.save(response)
        return response
