import json
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required, permission_required
from django.http import JsonResponse
from django.utils import timezone
from django.db.models import Sum, Q, F
from django.db import transaction

from sales.models import Invoice
from .models import Payment

class FinanceDashboardView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'finance.manage_finance'
    template_name = 'finance/dashboard.html'

    def get(self, request):
        today = timezone.now().date()
        
        from finance.stats import get_invoice_stats
        stats = get_invoice_stats(today)
        
        completed_qs = Invoice.objects.filter(status=Invoice.Status.PAID)
        pending_qs = Invoice.objects.filter(status=Invoice.Status.ISSUED)
        
        completed_total = completed_qs.aggregate(t=Sum('total_amount'))['t'] or 0
        pending_total = pending_qs.aggregate(t=Sum('total_amount'))['t'] or 0
        
        from .models import CustomerCredit
        total_credits = CustomerCredit.objects.filter(remaining_amount__gt=0, is_active=True).aggregate(t=Sum('remaining_amount'))['t'] or 0

        pending_reconciliation_count = Payment.objects.filter(
            reconciliation_status=Payment.ReconciliationStatus.UNRECONCILED
        ).count()
        
        context = {
            'total_invoices_all': stats['total_invoices_all'],
            'total_invoices_month': stats['total_invoices_month'],
            'completed_count': stats['completed_count'],
            'completed_total': completed_total,
            'pending_count': stats['pending_count'],
            'pending_total': pending_total,
            'partial_count': stats['partial_count'],
            'total_credits': total_credits,
            'pending_reconciliation_count': pending_reconciliation_count,
        }
        return render(request, self.template_name, context)

class PendingPaymentsView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'finance.manage_finance'
    template_name = 'finance/pending_payments.html'

    def get(self, request):
        today = timezone.now().date()
        
        valid_statuses = [Invoice.Status.ISSUED, Invoice.Status.PAID, Invoice.Status.EDIT_PENDING, Invoice.Status.CANCEL_PENDING]
        pending_qs = Invoice.objects.filter(status__in=valid_statuses).order_by('-creation_date')
        
        # Filter on creation_date so date range matches when the invoice was created
        date_from = request.GET.get('date_from')
        if date_from:
            pending_qs = pending_qs.filter(creation_date__date__gte=date_from)
            
        date_to = request.GET.get('date_to')
        if date_to:
            pending_qs = pending_qs.filter(creation_date__date__lte=date_to)
            
        month = request.GET.get('month')
        if month:
            try:
                year_str, month_str = month.split('-')
                pending_qs = pending_qs.filter(creation_date__year=int(year_str), creation_date__month=int(month_str))
            except ValueError:
                pass
            
        q = request.GET.get('q')
        if q:
            pending_qs = pending_qs.filter(
                Q(invoice_number__icontains=q) |
                Q(customer__customer_name__icontains=q)
            )

        salesperson_id = request.GET.get('salesperson')
        if salesperson_id:
            pending_qs = pending_qs.filter(salesperson_id=salesperson_id)

        from decimal import Decimal
        from django.db.models import Sum
        from django.db.models.functions import Coalesce
        
        # Annotate total paid at DB level, then filter — so pagination works correctly
        pending_qs = pending_qs.annotate(
            total_paid_sum=Coalesce(Sum('payments__amount'), Decimal('0.00'))
        ).filter(total_paid_sum__lt=Decimal('0.01'))  # treat < 0.01 as zero (tolerance)

        if request.GET.get('export') == 'xlsx':
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from django.http import HttpResponse
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Pending Payments"
            
            ws.append([f"Pending Payments Report (Generated {timezone.now().strftime('%Y-%m-%d %H:%M')})"])
            ws.append([])
            
            headers = ["Invoice #", "Type", "Customer", "Salesperson", "Creation Date", "Due Date", "Total (Rs)", "Total Paid (Rs)", "Balance (Rs)"]
            ws.append(headers)
            
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=3, column=col)
                cell.font = header_font
                cell.fill = header_fill
                
            for inv in pending_qs:
                ws.append([
                    inv.invoice_number,
                    inv.effective_payment_term_display,
                    inv.customer.customer_name if inv.customer else "-",
                    inv.salesperson.get_full_name() if inv.salesperson else "-",
                    inv.creation_date.strftime('%Y-%m-%d') if inv.creation_date else "-",
                    inv.due_date.strftime('%Y-%m-%d') if inv.due_date else "-",
                    float(inv.total_amount),
                    float(inv.total_paid_sum),
                    float(inv.total_amount - inv.total_paid_sum)
                ])
                
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="Pending_Payments.xlsx"'
            wb.save(response)
            return response

        from django.core.paginator import Paginator
        paginator = Paginator(pending_qs, 20)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)

        from .models import CustomerCredit
        from django.db.models import Sum

        customer_ids = [inv.customer.id for inv in page_obj.object_list if inv.customer]
        credits_qs = CustomerCredit.objects.filter(
            customer_id__in=customer_ids, remaining_amount__gt=0, is_active=True
        ).values('customer_id').annotate(total_credit=Sum('remaining_amount'))
        
        credit_map = {item['customer_id']: item['total_credit'] for item in credits_qs}

        # Build display list (no Python-level filtering needed now)
        invoices = []
        for inv in page_obj.object_list:
            total_paid = inv.total_paid_sum
            balance = inv.total_amount - total_paid
            avail_credit = credit_map.get(inv.customer.id if inv.customer else 0, Decimal('0.00'))
            invoices.append({
                    'id': inv.id,
                    'invoice_number': inv.invoice_number,
                    'invoice_type': inv.invoice_type,
                    'customer': inv.customer,
                    'salesperson': inv.salesperson,
                    'due_date': inv.due_date,
                    'total_amount': inv.total_amount,
                    'total_paid': total_paid,
                    'balance': balance,
                    'days_overdue': (today - inv.due_date).days if inv.due_date else 0,
                    'available_credit': avail_credit
                })
                
        from users.models import User
        sales_officers = User.objects.filter(role__name='Sales Officer', is_active=True).distinct()
        
        from finance.stats import get_invoice_stats
        stats = get_invoice_stats(today)
                
        context = {
            'invoices': invoices,
            'page_obj': page_obj,
            'paginator': paginator,
            'is_paginated': paginator.num_pages > 1,
            'payment_methods': Payment.PaymentMethod.choices,
            'bank_accounts': BankAccount.objects.filter(is_active=True),
            'sales_officers': sales_officers,
            'total_invoices_all': stats['total_invoices_all'],
            'total_invoices_month': stats['total_invoices_month'],
            'completed_count': stats['completed_count'],
            'pending_count': stats['pending_count'],
            'partial_count': stats['partial_count'],
        }
        
        # Provide Saved Filters
        try:
            from users.models import SavedFilter
            context['saved_filters'] = SavedFilter.objects.filter(user=request.user, model_name='PendingPayments')
        except ImportError:
            context['saved_filters'] = []
            
        return render(request, self.template_name, context)

class PartialPaymentsView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'finance.manage_finance'
    template_name = 'finance/partial_payments.html'

    def get(self, request):
        today = timezone.now().date()
        
        valid_statuses = [Invoice.Status.ISSUED, Invoice.Status.PAID, Invoice.Status.EDIT_PENDING, Invoice.Status.CANCEL_PENDING]
        pending_qs = Invoice.objects.filter(status__in=valid_statuses).order_by('due_date')
        
        # Filter on creation_date so date range matches when the invoice was created
        date_from = request.GET.get('date_from')
        if date_from:
            pending_qs = pending_qs.filter(creation_date__date__gte=date_from)
            
        date_to = request.GET.get('date_to')
        if date_to:
            pending_qs = pending_qs.filter(creation_date__date__lte=date_to)
            
        month = request.GET.get('month')
        if month:
            try:
                year_str, month_str = month.split('-')
                pending_qs = pending_qs.filter(creation_date__year=int(year_str), creation_date__month=int(month_str))
            except ValueError:
                pass
            
        q = request.GET.get('q')
        if q:
            pending_qs = pending_qs.filter(
                Q(invoice_number__icontains=q) |
                Q(customer__customer_name__icontains=q)
            )

        salesperson_id = request.GET.get('salesperson')
        if salesperson_id:
            pending_qs = pending_qs.filter(salesperson_id=salesperson_id)

        from decimal import Decimal
        from django.db.models import Sum, F
        from django.db.models.functions import Coalesce

        # Annotate total paid at DB level, then filter — so pagination works correctly
        # Use a 0.01 tolerance to match the same rounding used when recording payments
        pending_qs = pending_qs.annotate(
            total_paid_sum=Coalesce(Sum('payments__amount'), Decimal('0.00'))
        ).filter(total_paid_sum__gte=Decimal('0.01'), total_paid_sum__lt=F('total_amount') - Decimal('0.009'))

        if request.GET.get('export') == 'xlsx':
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from django.http import HttpResponse
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Partial Payments"
            
            ws.append([f"Partial Payments Report (Generated {timezone.now().strftime('%Y-%m-%d %H:%M')})"])
            ws.append([])
            
            headers = ["Invoice #", "Type", "Customer", "Salesperson", "Creation Date", "Due Date", "Total (Rs)", "Total Paid (Rs)", "Balance (Rs)"]
            ws.append(headers)
            
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=3, column=col)
                cell.font = header_font
                cell.fill = header_fill
                
            for inv in pending_qs:
                ws.append([
                    inv.invoice_number,
                    inv.effective_payment_term_display,
                    inv.customer.customer_name if inv.customer else "-",
                    inv.salesperson.get_full_name() if inv.salesperson else "-",
                    inv.creation_date.strftime('%Y-%m-%d') if inv.creation_date else "-",
                    inv.due_date.strftime('%Y-%m-%d') if inv.due_date else "-",
                    float(inv.total_amount),
                    float(inv.total_paid_sum),
                    float(inv.total_amount - inv.total_paid_sum)
                ])
                
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="Partial_Payments.xlsx"'
            wb.save(response)
            return response

        from django.core.paginator import Paginator
        paginator = Paginator(pending_qs, 20)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
        
        from .models import CustomerCredit
        from django.db.models import Sum

        customer_ids = [inv.customer.id for inv in page_obj.object_list if inv.customer]
        credits_qs = CustomerCredit.objects.filter(
            customer_id__in=customer_ids, remaining_amount__gt=0, is_active=True
        ).values('customer_id').annotate(total_credit=Sum('remaining_amount'))
        
        credit_map = {item['customer_id']: item['total_credit'] for item in credits_qs}

        # Build display list - prefetch payments to avoid N+1
        invoices = []
        for inv in page_obj.object_list.select_related('customer', 'salesperson').prefetch_related('payments'):
            total_paid = inv.total_paid_sum
            balance = inv.total_amount - total_paid
            payments = list(inv.payments.all().order_by('payment_date'))
            avail_credit = credit_map.get(inv.customer.id if inv.customer else 0, Decimal('0.00'))
            
            invoices.append({
                    'id': inv.id,
                    'invoice_number': inv.invoice_number,
                    'invoice_type': inv.invoice_type,
                    'customer': inv.customer,
                    'salesperson': inv.salesperson,
                    'due_date': inv.due_date,
                    'total_amount': inv.total_amount,
                    'total_paid': total_paid,
                    'balance': balance,
                    'days_overdue': (today - inv.due_date).days if inv.due_date else 0,
                    'payment_history': payments,
                    'available_credit': avail_credit
                })
        from users.models import User
        sales_officers = User.objects.filter(role__name='Sales Officer', is_active=True).distinct()
        
        from finance.stats import get_invoice_stats
        stats = get_invoice_stats(today)
                
        context = {
            'invoices': invoices,
            'page_obj': page_obj,
            'paginator': paginator,
            'is_paginated': paginator.num_pages > 1,
            'payment_methods': Payment.PaymentMethod.choices,
            'bank_accounts': BankAccount.objects.filter(is_active=True),
            'sales_officers': sales_officers,
            'total_invoices_all': stats['total_invoices_all'],
            'total_invoices_month': stats['total_invoices_month'],
            'completed_count': stats['completed_count'],
            'pending_count': stats['pending_count'],
            'partial_count': stats['partial_count'],
        }
        
        # Provide Saved Filters
        try:
            from users.models import SavedFilter
            context['saved_filters'] = SavedFilter.objects.filter(user=request.user, model_name='PartialPayments')
        except ImportError:
            context['saved_filters'] = []
            
        return render(request, self.template_name, context)

class CompletedPaymentsView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'finance.manage_finance'
    template_name = 'finance/completed_payments.html'

    def get(self, request):
        today = timezone.now().date()
        
        valid_statuses = [Invoice.Status.ISSUED, Invoice.Status.PAID, Invoice.Status.EDIT_PENDING, Invoice.Status.CANCEL_PENDING]
        from django.db.models import Sum, F
        from django.db.models.functions import Coalesce
        from decimal import Decimal
        completed_qs = Invoice.objects.filter(status__in=valid_statuses).annotate(
            total_paid_sum=Coalesce(Sum('payments__amount'), Decimal('0.00'))
        ).filter(total_paid_sum__gte=F('total_amount') - Decimal('0.009')).order_by('-creation_date')
        
        # Filter on creation_date so date range matches when the invoice was created
        date_from = request.GET.get('date_from')
        if date_from:
            completed_qs = completed_qs.filter(creation_date__date__gte=date_from)
            
        date_to = request.GET.get('date_to')
        if date_to:
            completed_qs = completed_qs.filter(creation_date__date__lte=date_to)
            
        month = request.GET.get('month')
        if month:
            try:
                year_str, month_str = month.split('-')
                completed_qs = completed_qs.filter(creation_date__year=int(year_str), creation_date__month=int(month_str))
            except ValueError:
                pass
            
        q = request.GET.get('q')
        if q:
            completed_qs = completed_qs.filter(
                Q(invoice_number__icontains=q) |
                Q(customer__customer_name__icontains=q)
            )

        salesperson_id = request.GET.get('salesperson')
        if salesperson_id:
            completed_qs = completed_qs.filter(salesperson_id=salesperson_id)

        if request.GET.get('export') == 'xlsx':
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from django.http import HttpResponse
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Completed Payments"
            
            ws.append([f"Completed Payments Report (Generated {timezone.now().strftime('%Y-%m-%d %H:%M')})"])
            ws.append([])
            
            headers = ["Invoice #", "Type", "Customer", "Salesperson", "Creation Date", "Due Date", "Total (Rs)", "Total Paid (Rs)", "Balance (Rs)"]
            ws.append(headers)
            
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=3, column=col)
                cell.font = header_font
                cell.fill = header_fill
                
            for inv in completed_qs:
                ws.append([
                    inv.invoice_number,
                    inv.effective_payment_term_display,
                    inv.customer.customer_name if inv.customer else "-",
                    inv.salesperson.get_full_name() if inv.salesperson else "-",
                    inv.creation_date.strftime('%Y-%m-%d') if inv.creation_date else "-",
                    inv.due_date.strftime('%Y-%m-%d') if inv.due_date else "-",
                    float(inv.total_amount),
                    float(inv.total_paid_sum),
                    float(inv.total_amount - inv.total_paid_sum)
                ])
                
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="Completed_Payments.xlsx"'
            wb.save(response)
            return response

        from django.core.paginator import Paginator
        paginator = Paginator(completed_qs, 20)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)

        # For completed invoices, total_paid = total_amount (they are PAID status)
        # Avoid N+1 by not querying payments per invoice here
        # Add prefetch for payment detail in history view instead
        invoices = []
        for inv in page_obj.object_list.select_related('customer', 'salesperson').prefetch_related('payments'):
            payments = list(inv.payments.all())
            total_paid = sum(p.amount for p in payments)
            last_payment = payments[-1] if payments else None
            invoices.append({
                'id': inv.id,
                'invoice_number': inv.invoice_number,
                'invoice_type': inv.invoice_type,
                'customer': inv.customer,
                'salesperson': inv.salesperson,
                'due_date': inv.due_date,
                'total_amount': inv.total_amount,
                'total_paid': total_paid,
                'balance': 0,
                'last_payment': last_payment,
                'payment_count': len(payments),
                'payment_history': payments,
            })
                
        from users.models import User
        sales_officers = User.objects.filter(role__name='Sales Officer', is_active=True).distinct()
        
        from finance.stats import get_invoice_stats
        stats = get_invoice_stats(today)
                
        context = {
            'invoices': invoices,
            'page_obj': page_obj,
            'paginator': paginator,
            'is_paginated': paginator.num_pages > 1,
            'sales_officers': sales_officers,
            'total_invoices_all': stats['total_invoices_all'],
            'total_invoices_month': stats['total_invoices_month'],
            'completed_count': stats['completed_count'],
            'pending_count': stats['pending_count'],
            'partial_count': stats['partial_count'],
        }
        
        # Provide Saved Filters
        try:
            from users.models import SavedFilter
            context['saved_filters'] = SavedFilter.objects.filter(user=request.user, model_name='CompletedPayments')
        except ImportError:
            context['saved_filters'] = []
            
        return render(request, self.template_name, context)

class RecordPaymentView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'finance.manage_finance'

    def post(self, request):
        try:
            # Using request.POST and request.FILES since the frontend now sends FormData
            data = request.POST
            invoice_id = data.get('invoice_id')
            amount_str = data.get('amount')
            payment_date_str = data.get('payment_date')
            payment_method = data.get('payment_method')
            bank_account_id = data.get('bank_account_id')
            reference = data.get('reference', '').strip()
            notes = data.get('notes', '').strip()
            slip = request.FILES.get('slip_attachment')

            if not amount_str:
                raise ValueError("Payment amount is required.")
                
            bank_account = None
            if payment_method in ['BANK_TRANSFER', 'CHEQUE', 'CARD']:
                if not bank_account_id:
                    raise ValueError("You must select a bank account for this payment method.")
                from .models import BankAccount
                bank_account = get_object_or_404(BankAccount, id=bank_account_id)

            invoice = get_object_or_404(Invoice, id=invoice_id)
            from decimal import Decimal
            amount = Decimal(amount_str)
            
            if amount <= 0:
                raise ValueError("Payment amount must be greater than 0.")
                
            total_paid = sum(p.amount for p in invoice.payments.all())
            balance = invoice.total_amount - total_paid
            
            payment_amount_to_invoice = amount
            credit_amount = Decimal('0.00')

            if amount > balance + Decimal('0.01'):
                if balance > Decimal('0.00'):
                    payment_amount_to_invoice = balance
                else:
                    payment_amount_to_invoice = Decimal('0.00')
                credit_amount = amount - payment_amount_to_invoice
                
            if not reference and not slip:
                raise ValueError("You must provide either a reference number or upload a payment slip.")

            payment_date = datetime.strptime(payment_date_str, '%Y-%m-%d').date()

            with transaction.atomic():
                if payment_amount_to_invoice > 0:
                    Payment.objects.create(
                        invoice=invoice,
                        amount=payment_amount_to_invoice,
                        payment_date=payment_date,
                        payment_method=payment_method,
                        bank_account=bank_account,
                        reference_number=reference,
                        slip_attachment=slip,
                        notes=notes,
                        recorded_by=request.user
                    )
                
                if credit_amount > 0:
                    from .models import CustomerCredit
                    if not invoice.customer:
                        raise ValueError("Cannot create an overpayment credit because this invoice has no linked customer.")
                    CustomerCredit.objects.create(
                        customer=invoice.customer,
                        original_amount=credit_amount,
                        remaining_amount=credit_amount,
                        notes=f"Overpayment on Invoice {invoice.invoice_number} (Ref: {reference})"
                    )
                
                # Create Journal Entry for the Payment
                from .models import JournalEntry, JournalEntryLine, Account
                je = JournalEntry.objects.create(
                    date=payment_date,
                    reference=f"Payment for {invoice.invoice_number} (Ref: {reference})",
                    created_by=request.user,
                    status=JournalEntry.Status.POSTED
                )
                
                # Debit Account
                if bank_account:
                    debit_account = bank_account.ledger_account
                else:
                    debit_account = Account.objects.get(code='1000') # Cash
                    
                JournalEntryLine.objects.create(
                    journal_entry=je,
                    account=debit_account,
                    description=f"Received payment for {invoice.invoice_number}",
                    debit=amount,
                    credit=0
                )
                
                # Credit Accounts Receivable
                ar_account = Account.objects.get(code='1200')
                JournalEntryLine.objects.create(
                    journal_entry=je,
                    account=ar_account,
                    description=f"Payment applied to {invoice.invoice_number}",
                    debit=0,
                    credit=amount
                )

            return JsonResponse({'status': 'ok', 'message': 'Payment recorded successfully.'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
class AgedReceivablesView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'finance.manage_finance'
    template_name = 'finance/aged_receivables.html'

    def get(self, request):
        today = timezone.now().date()
        as_of_date_str = request.GET.get('as_of_date')
        if as_of_date_str:
            from datetime import datetime
            try:
                as_of_date = datetime.strptime(as_of_date_str, '%Y-%m-%d').date()
            except ValueError:
                as_of_date = today
        else:
            as_of_date = today
            
        qs = Invoice.objects.filter(
            status__in=[Invoice.Status.ISSUED, Invoice.Status.PAID],
            creation_date__date__lte=as_of_date
        ).exclude(invoice_type__in=['CASH', 'COD']).select_related('customer', 'salesperson').prefetch_related('payments')
        
        q = request.GET.get('q')
        if q:
            from django.db.models import Q
            qs = qs.filter(
                Q(invoice_number__icontains=q) |
                Q(customer__customer_name__icontains=q)
            )
            
        salesperson_id = request.GET.get('salesperson')
        if salesperson_id:
            qs = qs.filter(salesperson_id=salesperson_id)
        
        customers_data = {}
        
        for inv in qs:
            total_paid = sum(p.amount for p in inv.payments.all() if p.payment_date <= as_of_date)
            balance = inv.total_amount - total_paid
            
            if balance <= 0.005:
                continue
                
            customer_id = inv.customer.id
            if customer_id not in customers_data:
                customers_data[customer_id] = {
                    'customer': inv.customer,
                    'invoices': [],
                    'not_due': 0,
                    '1_30': 0,
                    '31_45': 0,
                    '46_60': 0,
                    '61_90': 0,
                    '91_120': 0,
                    'older': 0,
                    'total': 0
                }
                
            days_overdue = (as_of_date - inv.due_date).days if inv.due_date else 0
            
            inv_data = {
                'id': inv.id,
                'invoice_number': inv.invoice_number,
                'salesperson': inv.salesperson,
                'due_date': inv.due_date,
                'total_amount': float(inv.total_amount),
                'balance': float(balance),
                'days_overdue': days_overdue
            }
            
            if days_overdue <= 0:
                customers_data[customer_id]['not_due'] += float(balance)
                inv_data['bucket'] = 'not_due'
            elif 1 <= days_overdue <= 30:
                customers_data[customer_id]['1_30'] += float(balance)
                inv_data['bucket'] = '1_30'
            elif 31 <= days_overdue <= 45:
                customers_data[customer_id]['31_45'] += float(balance)
                inv_data['bucket'] = '31_45'
            elif 46 <= days_overdue <= 60:
                customers_data[customer_id]['46_60'] += float(balance)
                inv_data['bucket'] = '46_60'
            elif 61 <= days_overdue <= 90:
                customers_data[customer_id]['61_90'] += float(balance)
                inv_data['bucket'] = '61_90'
            elif 91 <= days_overdue <= 120:
                customers_data[customer_id]['91_120'] += float(balance)
                inv_data['bucket'] = '91_120'
            else:
                customers_data[customer_id]['older'] += float(balance)
                inv_data['bucket'] = 'older'
                
            customers_data[customer_id]['total'] += float(balance)
            customers_data[customer_id]['invoices'].append(inv_data)
            
        customers_list = list(customers_data.values())
        customers_list.sort(key=lambda x: x['total'], reverse=True)
        
        grand_totals = {
            'not_due': sum(c['not_due'] for c in customers_list),
            '1_30': sum(c['1_30'] for c in customers_list),
            '31_45': sum(c['31_45'] for c in customers_list),
            '46_60': sum(c['46_60'] for c in customers_list),
            '61_90': sum(c['61_90'] for c in customers_list),
            '91_120': sum(c['91_120'] for c in customers_list),
            'older': sum(c['older'] for c in customers_list),
            'total': sum(c['total'] for c in customers_list),
        }
        
        export_format = request.GET.get('export')
        if export_format == 'pdf':
            context = {
                'customers_data': customers_list,
                'grand_totals': grand_totals,
                'as_of_date': as_of_date
            }
            return render(request, 'finance/aged_receivables_print.html', context)
            
        if export_format == 'xlsx':
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from django.http import HttpResponse
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Aged Receivables"
            
            ws.append([f"Aged Receivables Report (As of {as_of_date.strftime('%Y-%m-%d')})"])
            ws.append([])
            
            headers = ["Customer / Invoice", "Salesperson", "Expected Date", "Not Due", "1-30", "31-45", "46-60", "61-90", "91-120", "Older", "Total"]
            ws.append(headers)
            
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
            
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=3, column=col)
                cell.font = header_font
                cell.fill = header_fill
                
            for c in customers_list:
                row = [
                    c['customer'].customer_name,
                    "",
                    "",
                    c['not_due'],
                    c['1_30'],
                    c['31_45'],
                    c['46_60'],
                    c['61_90'],
                    c['91_120'],
                    c['older'],
                    c['total']
                ]
                ws.append(row)
                for i in range(1, len(headers)+1):
                    ws.cell(row=ws.max_row, column=i).font = Font(bold=True)
                
                for inv in c['invoices']:
                    ws.append([
                        inv['invoice_number'],
                        inv['salesperson'].get_full_name() if inv['salesperson'] else "",
                        inv['due_date'].strftime('%Y-%m-%d') if inv['due_date'] else "",
                        inv['balance'] if inv['bucket'] == 'not_due' else "",
                        inv['balance'] if inv['bucket'] == '1_30' else "",
                        inv['balance'] if inv['bucket'] == '31_45' else "",
                        inv['balance'] if inv['bucket'] == '46_60' else "",
                        inv['balance'] if inv['bucket'] == '61_90' else "",
                        inv['balance'] if inv['bucket'] == '91_120' else "",
                        inv['balance'] if inv['bucket'] == 'older' else "",
                        inv['balance']
                    ])
            
            ws.append([])
            ws.append([
                "GRAND TOTAL", "", "",
                grand_totals['not_due'],
                grand_totals['1_30'],
                grand_totals['31_45'],
                grand_totals['46_60'],
                grand_totals['61_90'],
                grand_totals['91_120'],
                grand_totals['older'],
                grand_totals['total']
            ])
            for i in range(1, len(headers)+1):
                ws.cell(row=ws.max_row, column=i).font = Font(bold=True)
                
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="Aged_Receivables_{as_of_date.strftime("%Y%m%d")}.xlsx"'
            wb.save(response)
            return response
            
        from users.models import User
        sales_officers = User.objects.filter(role__name='Sales Officer', is_active=True).distinct()
        
        context = {
            'customers_data': customers_list,
            'grand_totals': grand_totals,
            'as_of_date': as_of_date,
            'sales_officers': sales_officers
        }
        
        return render(request, self.template_name, context)

from django.urls import reverse_lazy
from django.views.generic import View, TemplateView, ListView, CreateView, UpdateView, DeleteView, DetailView
from .models import Account, AccountType, JournalEntry, JournalEntryLine
from .forms import AccountForm, JournalEntryForm

class ChartOfAccountsView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = 'finance.manage_finance'
    template_name = 'finance/chart_of_accounts.html'
    model = Account
    context_object_name = 'accounts'

    def get(self, request, *args, **kwargs):
        if request.GET.get('export') == 'xlsx':
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from django.http import HttpResponse
            from django.utils import timezone
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Chart of Accounts"
            ws.append([f"Chart of Accounts Report (Generated {timezone.now().strftime('%Y-%m-%d %H:%M')})"])
            ws.append([])
            
            headers = ["Code", "Account Name", "Type", "Sub Type", "Balance (Rs)", "Status"]
            ws.append(headers)
            
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=3, column=col)
                cell.font = header_font
                cell.fill = header_fill
                
            for acc in self.model.objects.all().order_by('code'):
                ws.append([
                    acc.code,
                    acc.name,
                    acc.get_account_type_display(),
                    acc.get_sub_type_display(),
                    float(acc.balance),
                    "Active" if acc.is_active else "Inactive"
                ])
                
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="Chart_Of_Accounts.xlsx"'
            wb.save(response)
            return response
            
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Group accounts by type for the view
        grouped = {}
        for account_type in AccountType.choices:
            grouped[account_type[1]] = Account.objects.filter(account_type=account_type[0])
        context['grouped_accounts'] = grouped
        return context

class AccountCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'finance.manage_finance'
    template_name = 'finance/account_form.html'
    form_class = AccountForm
    success_url = reverse_lazy('finance_chart_of_accounts')

class JournalEntryListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = 'finance.manage_finance'
    template_name = 'finance/journal_entries.html'
    model = JournalEntry
    context_object_name = 'entries'
    paginate_by = 50
    
    def get_queryset(self):
        from django.db.models import Sum
        return super().get_queryset().annotate(
            total_amount=Sum('lines__debit')
        )
    
    def get(self, request, *args, **kwargs):
        if request.GET.get('export') == 'xlsx':
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from django.http import HttpResponse
            from django.utils import timezone
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Journal Entries"
            ws.append([f"Journal Entries Report (Generated {timezone.now().strftime('%Y-%m-%d %H:%M')})"])
            ws.append([])
            
            headers = ["JE #", "Date", "Reference", "Status", "Account", "Description", "Debit (Rs)", "Credit (Rs)"]
            ws.append(headers)
            
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=3, column=col)
                cell.font = header_font
                cell.fill = header_fill
                
            entries = self.model.objects.all().prefetch_related('lines__account').order_by('-date', '-id')
            for entry in entries:
                for line in entry.lines.all():
                    ws.append([
                        f"JE-{entry.id}",
                        entry.date.strftime('%Y-%m-%d') if entry.date else "",
                        entry.reference or "-",
                        entry.get_status_display(),
                        line.account.name if line.account else "-",
                        line.description or "-",
                        float(line.debit) if line.debit else 0.0,
                        float(line.credit) if line.credit else 0.0
                    ])
                
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="Journal_Entries.xlsx"'
            wb.save(response)
            return response
            
        return super().get(request, *args, **kwargs)

class JournalEntryDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    permission_required = 'finance.manage_finance'
    template_name = 'finance/journal_entry_detail.html'
    model = JournalEntry
    context_object_name = 'entry'

    def get_queryset(self):
        from django.db.models import Sum
        return super().get_queryset().prefetch_related('lines__account').annotate(
            total_amount=Sum('lines__debit')
        )
class JournalEntryCreateView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'finance.manage_finance'
    template_name = 'finance/journal_entry_form.html'

    def get(self, request):
        accounts = Account.objects.filter(is_active=True)
        return render(request, self.template_name, {'accounts': accounts})

    def post(self, request):
        try:
            data = json.loads(request.body)
            date_str = data.get('date')
            reference = data.get('reference', '')
            lines = data.get('lines', [])
            
            # Validation
            if not lines or len(lines) < 2:
                return JsonResponse({'status': 'error', 'message': 'At least two lines are required.'}, status=400)
                
            total_debit = sum(float(l.get('debit') or 0) for l in lines)
            total_credit = sum(float(l.get('credit') or 0) for l in lines)
            
            if abs(total_debit - total_credit) > 0.01:
                return JsonResponse({'status': 'error', 'message': f'Debits ({total_debit}) and Credits ({total_credit}) must balance.'}, status=400)

            with transaction.atomic():
                je = JournalEntry.objects.create(
                    date=datetime.strptime(date_str, '%Y-%m-%d').date(),
                    reference=reference,
                    created_by=request.user,
                    status=JournalEntry.Status.DRAFT
                )
                
                # Check permissions if trying to post
                if data.get('action') == 'POST':
                    if request.user.has_perm('finance.post_journal_entry'):
                        je.status = JournalEntry.Status.POSTED
                        je.save()
                    else:
                        return JsonResponse({'status': 'error', 'message': 'You do not have permission to post journal entries. Save as Draft instead.'}, status=403)
                
                for line in lines:
                    JournalEntryLine.objects.create(
                        journal_entry=je,
                        account_id=line['account_id'],
                        description=line.get('description', ''),
                        debit=float(line.get('debit') or 0),
                        credit=float(line.get('credit') or 0)
                    )
            
            return JsonResponse({'status': 'ok', 'message': 'Journal Entry created successfully.'})
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

class GeneralLedgerView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'finance.manage_finance'
    template_name = 'finance/general_ledger.html'

    def get(self, request):
        accounts = Account.objects.filter(is_active=True)
        
        account_id = request.GET.get('account')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        
        lines = []
        selected_account = None
        running_balance = 0
        
        if account_id:
            selected_account = get_object_or_404(Account, id=account_id)
            qs = JournalEntryLine.objects.filter(
                account=selected_account,
                journal_entry__status=JournalEntry.Status.POSTED
            ).select_related('journal_entry')
            
            if date_from:
                qs = qs.filter(journal_entry__date__gte=date_from)
            if date_to:
                qs = qs.filter(journal_entry__date__lte=date_to)
                
            qs = qs.order_by('journal_entry__date', 'journal_entry__id')
            
            for line in qs:
                # Calculate balance depending on account type
                # Asset/Expense increase with Debit. Liability/Equity/Revenue increase with Credit.
                amount = line.debit - line.credit if selected_account.account_type in [AccountType.ASSET, AccountType.EXPENSE] else line.credit - line.debit
                running_balance += amount
                
                lines.append({
                    'date': line.journal_entry.date,
                    'reference': line.journal_entry.reference,
                    'description': line.description,
                    'debit': line.debit,
                    'credit': line.credit,
                    'balance': running_balance
                })

        if request.GET.get('export') == 'xlsx':
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from django.http import HttpResponse
            from django.utils import timezone
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "General Ledger"
            
            acc_name = selected_account.name if selected_account else 'All Accounts'
            ws.append([f"General Ledger Report - {acc_name} (Generated {timezone.now().strftime('%Y-%m-%d %H:%M')})"])
            ws.append([])
            
            headers = ["Date", "Reference", "Description", "Debit (Rs)", "Credit (Rs)", "Balance (Rs)"]
            ws.append(headers)
            
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=3, column=col)
                cell.font = header_font
                cell.fill = header_fill
                
            for line in lines:
                ws.append([
                    line['date'].strftime('%Y-%m-%d') if line['date'] else "",
                    line['reference'] or "-",
                    line['description'] or "-",
                    float(line['debit']),
                    float(line['credit']),
                    float(line['balance'])
                ])
                
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="General_Ledger.xlsx"'
            wb.save(response)
            return response
                
        context = {
            'accounts': accounts,
            'selected_account': selected_account,
            'lines': lines
        }
        return render(request, self.template_name, context)

class CustomerCreditListView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'finance.manage_finance'
    template_name = 'finance/customer_credit_list.html'

    def get(self, request):
        from .models import CustomerCredit
        credits = CustomerCredit.objects.filter(remaining_amount__gt=0, is_active=True).select_related('customer').order_by('-created_at')
        
        from django.db.models import Sum
        customer_totals = credits.values('customer__customer_name', 'customer_id').annotate(total=Sum('remaining_amount')).order_by('-total')
        
        context = {
            'credits': credits,
            'customer_totals': customer_totals
        }
        return render(request, self.template_name, context)

@login_required
@permission_required('finance.manage_finance', raise_exception=True)
def apply_customer_credit(request):
    if request.method == 'POST':
        from .models import CustomerCredit, CreditApplication, Payment
        from sales.models import Invoice
        from decimal import Decimal
        
        invoice_id = request.POST.get('invoice_id')
        amount_str = request.POST.get('amount')
        
        try:
            invoice = Invoice.objects.get(id=invoice_id)
            amount = Decimal(amount_str)
            
            if amount <= 0:
                raise ValueError("Amount must be greater than 0.")
            
            from django.db.models import Sum
            credits = CustomerCredit.objects.filter(customer=invoice.customer, remaining_amount__gt=0, is_active=True).order_by('created_at')
            total_available = credits.aggregate(t=Sum('remaining_amount'))['t'] or Decimal('0.00')
            
            if amount > total_available:
                raise ValueError(f"Cannot apply more than the available credit balance (Rs {total_available:.2f}).")
                
            total_paid = sum(p.amount for p in invoice.payments.all())
            balance = invoice.total_amount - total_paid
            
            if amount > balance + Decimal('0.01'):
                raise ValueError("Cannot apply more than the invoice balance.")
                
            with transaction.atomic():
                remaining_to_apply = amount
                for credit in credits:
                    if remaining_to_apply <= 0:
                        break
                    
                    apply_amt = min(credit.remaining_amount, remaining_to_apply)
                    credit.remaining_amount -= apply_amt
                    credit.save()
                    
                    CreditApplication.objects.create(
                        customer_credit=credit,
                        invoice=invoice,
                        amount_applied=apply_amt,
                        applied_by=request.user
                    )
                    
                    remaining_to_apply -= apply_amt
                
                Payment.objects.create(
                    invoice=invoice,
                    amount=amount,
                    payment_date=timezone.now().date(),
                    payment_method=Payment.PaymentMethod.OTHER,
                    reference_number=f"Credit Application",
                    notes=f"Applied from Customer Credit",
                    recorded_by=request.user
                )
            
            return JsonResponse({'status': 'ok', 'message': 'Credit applied successfully.'})
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'}, status=405)


# ─── Payment Reconciliation Views ────────────────────────────────────────────

class ReconciliationView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Displays all unreconciled payments for review and confirmation."""
    permission_required = 'finance.manage_finance'
    template_name = 'finance/reconciliation.html'

    def get(self, request):
        from decimal import Decimal

        qs = Payment.objects.select_related(
            'invoice', 'invoice__customer', 'recorded_by'
        ).filter(
            reconciliation_status=Payment.ReconciliationStatus.UNRECONCILED
        ).order_by('-payment_date', '-created_at')

        # Filters
        date_from = request.GET.get('date_from')
        if date_from:
            qs = qs.filter(payment_date__gte=date_from)

        date_to = request.GET.get('date_to')
        if date_to:
            qs = qs.filter(payment_date__lte=date_to)

        q = request.GET.get('q', '').strip()
        if q:
            qs = qs.filter(
                Q(invoice__invoice_number__icontains=q) |
                Q(invoice__customer__customer_name__icontains=q) |
                Q(reference_number__icontains=q)
            )

        method = request.GET.get('method')
        if method:
            qs = qs.filter(payment_method=method)
            
        bank_id = request.GET.get('bank_account_id')
        if bank_id:
            qs = qs.filter(bank_account_id=bank_id)

        total_amount = qs.aggregate(t=Sum('amount'))['t'] or Decimal('0.00')
        count = qs.count()

        # Reconciled payments for the history tab
        reconciled_qs = Payment.objects.select_related(
            'invoice', 'invoice__customer', 'reconciled_by'
        ).filter(
            reconciliation_status=Payment.ReconciliationStatus.RECONCILED
        ).order_by('-reconciled_at')[:50]

        context = {
            'payments': qs,
            'reconciled_payments': reconciled_qs,
            'total_amount': total_amount,
            'count': count,
            'payment_methods': Payment.PaymentMethod.choices,
            'bank_accounts': BankAccount.objects.filter(is_active=True),
            'filters': {
                'date_from': date_from or '',
                'date_to': date_to or '',
                'q': q,
                'method': method or '',
            }
        }
        return render(request, self.template_name, context)


class ReconcilePaymentView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """API: Reconcile a single payment."""
    permission_required = 'finance.manage_finance'

    def post(self, request):
        try:
            data = json.loads(request.body)
            payment_id = data.get('payment_id')
            note = data.get('note', '').strip()
            action = data.get('action', 'RECONCILED')  # RECONCILED or DISPUTED

            payment = get_object_or_404(Payment, id=payment_id)

            if action not in [Payment.ReconciliationStatus.RECONCILED, Payment.ReconciliationStatus.DISPUTED]:
                return JsonResponse({'status': 'error', 'message': 'Invalid action.'}, status=400)

            payment.reconciliation_status = action
            payment.reconciled_by = request.user
            payment.reconciled_at = timezone.now()
            if note:
                payment.reconciliation_note = note
            payment.save(update_fields=['reconciliation_status', 'reconciled_by', 'reconciled_at', 'reconciliation_note'])

            return JsonResponse({
                'status': 'ok',
                'message': f'Payment marked as {action.lower()}.',
                'payment_id': payment.id,
                'new_status': action,
            })

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


class BulkReconcileView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """API: Reconcile multiple payments at once."""
    permission_required = 'finance.manage_finance'

    def post(self, request):
        try:
            data = json.loads(request.body)
            payment_ids = data.get('payment_ids', [])
            note = data.get('note', '').strip()

            if not payment_ids:
                return JsonResponse({'status': 'error', 'message': 'No payments selected.'}, status=400)

            now = timezone.now()
            updated = Payment.objects.filter(
                id__in=payment_ids,
                reconciliation_status=Payment.ReconciliationStatus.UNRECONCILED
            ).update(
                reconciliation_status=Payment.ReconciliationStatus.RECONCILED,
                reconciled_by=request.user,
                reconciled_at=now,
                reconciliation_note=note or None,
            )

            return JsonResponse({
                'status': 'ok',
                'message': f'{updated} payment(s) reconciled successfully.',
                'reconciled_count': updated,
            })

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

from .models import BankAccount
from .forms import BankAccountForm

class BankAccountListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = 'finance.manage_finance'
    template_name = 'finance/bank_account_list.html'
    model = BankAccount
    context_object_name = 'bank_accounts'

class BankAccountCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'finance.manage_finance'
    template_name = 'finance/bank_account_form.html'
    form_class = BankAccountForm
    success_url = reverse_lazy('finance_bank_accounts')

class BankAccountUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'finance.manage_finance'
    model = BankAccount
    template_name = 'finance/bank_account_form.html'
    form_class = BankAccountForm
    success_url = reverse_lazy('finance_bank_accounts')

class BankAccountDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = BankAccount
    success_url = reverse_lazy('finance_bank_accounts')
    
    def test_func(self):
        return self.request.user.is_superuser
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, "Bank account deleted successfully.")
        return super().delete(request, *args, **kwargs)
